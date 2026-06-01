from __future__ import annotations

import html
import json
import os
import re
import shutil
import time
import hashlib
import textwrap
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

import bcrypt
import jwt
import psycopg
from dotenv import load_dotenv
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydub import AudioSegment
from psycopg.types.json import Json
from faster_whisper import WhisperModel
from task_rule_engine import analyze_task_candidate
from qa_postprocessor import build_qa_pairs_global
from speaker_diarization import apply_speaker_diarization
TORCH_AVAILABLE = False
NATASHA_AVAILABLE = False
EXCEL_AVAILABLE = False
PDF_AVAILABLE = False

try:
    import torch
    import torch.nn.functional as F
    from transformers import AutoTokenizer, AutoModelForSequenceClassification, pipeline
    TORCH_AVAILABLE = True
except Exception:
    pass

try:
    from natasha import Segmenter, MorphVocab, NewsEmbedding, NewsNERTagger, DatesExtractor, Doc
    NATASHA_AVAILABLE = True
except Exception:
    pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except Exception as e:
    print(f"[init] excel disabled: {e}")
    EXCEL_AVAILABLE = False

try:
    from fpdf import FPDF
    from fpdf.errors import FPDFException
    PDF_AVAILABLE = True
except Exception as e:
    print(f"[init] pdf disabled: {e}")
    PDF_AVAILABLE = False

load_dotenv()
os.environ.setdefault("TOKENIZERS_PARALLELISM", "false")

DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL not found in .env")

WHISPER_MODEL_NAME = os.getenv("WHISPER_MODEL_NAME", "large-v3")
WHISPER_DEVICE = os.getenv("WHISPER_DEVICE", "cpu").lower()
if WHISPER_DEVICE not in {"cpu", "cuda"}:
    WHISPER_DEVICE = "cpu"

WHISPER_COMPUTE_TYPE = os.getenv(
    "WHISPER_COMPUTE_TYPE",
    "float16" if WHISPER_DEVICE == "cuda" else "int8",
)

CLASSIFIER_DIR = os.getenv("CLASSIFIER_DIR", "models/classifier")
SENTIMENT_MODEL_NAME = os.getenv("SENTIMENT_MODEL_NAME", "seara/rubert-tiny2-russian-sentiment")

# Порог не должен быть слишком высоким: на реальных встречах реплики шумные,
# поэтому модель должна иметь возможность отдавать задачи без жесткого rules-фильтра.
CLASSIFIER_CONFIDENCE = float(os.getenv("CLASSIFIER_CONFIDENCE_THRESHOLD", "0.60"))
CLASSIFIER_TASK_CONFIDENCE = float(os.getenv("CLASSIFIER_TASK_CONFIDENCE_THRESHOLD", "0.55"))
CLASSIFIER_QUESTION_CONFIDENCE = float(os.getenv("CLASSIFIER_QUESTION_CONFIDENCE_THRESHOLD", "0.60"))
CLASSIFIER_ANSWER_CONFIDENCE = float(os.getenv("CLASSIFIER_ANSWER_CONFIDENCE_THRESHOLD", "0.62"))
INFERENCE_BATCH_SIZE = int(os.getenv("INFERENCE_BATCH_SIZE", "32"))

# Дополнительное дробление Whisper-сегментов. Без этого длинный сегмент может
# содержать одновременно задачу, вопрос и мусор, и классификатор отдаст other.
SPLIT_TRANSCRIPT_SEGMENTS = os.getenv("SPLIT_TRANSCRIPT_SEGMENTS", "1").lower() not in {"0", "false", "no"}
MAX_SEGMENT_WORDS = int(os.getenv("MAX_SEGMENT_WORDS", "32"))
MIN_SEGMENT_CHARS = int(os.getenv("MIN_SEGMENT_CHARS", "6"))

WHISPER_LANGUAGE = os.getenv("WHISPER_LANGUAGE") or None
WHISPER_BEAM_SIZE = int(os.getenv("WHISPER_BEAM_SIZE", "3"))

QA_MAX_GAP_SECONDS = float(os.getenv("QA_MAX_GAP_SECONDS", "45"))
QA_CONTEXT_WINDOW_SECONDS = float(os.getenv("QA_CONTEXT_WINDOW_SECONDS", "120"))

QA_MIN_QUESTION_CHARS = int(os.getenv("QA_MIN_QUESTION_CHARS", "10"))
QA_MIN_ANSWER_CHARS = int(os.getenv("QA_MIN_ANSWER_CHARS", "12"))
QA_MAX_ANSWER_SEGMENTS = int(os.getenv("QA_MAX_ANSWER_SEGMENTS", "12"))

BASE_DIR = Path(".")
UPLOAD_DIR = BASE_DIR / "uploads"
CONVERTED_DIR = BASE_DIR / "converted"
RESULTS_DIR = BASE_DIR / "results"
EXPORTS_DIR = BASE_DIR / "exports"

for d in [UPLOAD_DIR, CONVERTED_DIR, RESULTS_DIR, EXPORTS_DIR]:
    d.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {".mp3", ".wav", ".m4a", ".opus"}

JWT_SECRET = os.getenv("JWT_SECRET", "change-me-in-production")
JWT_ALGORITHM = os.getenv("JWT_ALGORITHM", "HS256")
JWT_EXPIRE_MINUTES = int(os.getenv("JWT_EXPIRE_MINUTES", "1440"))

app = FastAPI(title="PM Insights")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/results", StaticFiles(directory=str(RESULTS_DIR)), name="results")
app.mount("/exports", StaticFiles(directory=str(EXPORTS_DIR)), name="exports")

WHISPER_FALLBACK_CHAIN = [WHISPER_MODEL_NAME]
_seen = set()
WHISPER_FALLBACK_CHAIN = [m for m in WHISPER_FALLBACK_CHAIN if not (m in _seen or _seen.add(m))]

whisper_model = None
for _model_name in WHISPER_FALLBACK_CHAIN:
    try:
        print(f"[whisper] loading model '{_model_name}' (device={WHISPER_DEVICE}, compute={WHISPER_COMPUTE_TYPE})")
        whisper_model = WhisperModel(
            _model_name,
            device=WHISPER_DEVICE,
            compute_type=WHISPER_COMPUTE_TYPE,
        )
        WHISPER_MODEL_NAME = _model_name
        print(f"[whisper] model '{_model_name}' loaded successfully")
        break
    except Exception as e:
        print(f"[whisper] failed to load '{_model_name}': {e}")

if whisper_model is None:
    raise RuntimeError("Could not load any Whisper model")

classifier_tokenizer = None
classifier_model = None
sentiment_pipe = None
segmenter = morph_vocab = emb_natasha = ner_tagger = dates_extractor = None

NLP_MODE = "rules-only"
LABELS = {0: "task", 1: "question", 2: "answer", 3: "other"}

if TORCH_AVAILABLE:
    cdir = Path(CLASSIFIER_DIR)
    if cdir.exists():
        try:
            classifier_tokenizer = AutoTokenizer.from_pretrained(str(cdir))
            classifier_model = AutoModelForSequenceClassification.from_pretrained(str(cdir))
            device = "cuda" if torch.cuda.is_available() else "cpu"
            classifier_model.to(device)
            classifier_model.eval()
            NLP_MODE = "classifier + rules"
            print(f"[classifier] loaded from {cdir.resolve()} on {device}")
        except Exception as e:
            print(f"[init] classifier disabled: {e}")
            classifier_tokenizer = None
            classifier_model = None

    try:
        sentiment_pipe = pipeline(
            "text-classification",
            model=SENTIMENT_MODEL_NAME,
            tokenizer=SENTIMENT_MODEL_NAME,
            device=0 if torch.cuda.is_available() else -1,
            truncation=True,
        )
        NLP_MODE += " + sentiment"
    except Exception as e:
        print(f"[init] sentiment disabled: {e}")
        sentiment_pipe = None

if NATASHA_AVAILABLE:
    try:
        segmenter = Segmenter()
        morph_vocab = MorphVocab()
        emb_natasha = NewsEmbedding()
        ner_tagger = NewsNERTagger(emb_natasha)
        dates_extractor = DatesExtractor(morph_vocab)
        NLP_MODE += " + natasha"
    except Exception as e:
        print(f"[init] natasha disabled: {e}")
        segmenter = morph_vocab = emb_natasha = ner_tagger = dates_extractor = None

FILLER_WORDS = {
    "ну", "вот", "как", "бы", "типа", "короче", "значит", "получается", "наверное", "наверно",
    "вообще", "просто", "собственно", "то", "есть", "там", "тут", "это", "этот", "эта", "эти",
    "какой-то", "какая-то", "какие-то", "что-то", "как-то", "где-то", "может", "может быть",
}

INTERROGATIVE_WORDS = (
    "кто", "что", "где", "когда", "почему", "зачем", "как",
    "какой", "какая", "какое", "какие", "сколько",
    "можно ли", "нужно ли", "успеем ли", "в чём", "в чем",
    "а что", "что за", "можешь", "можете", "можем",
    "расскажи", "поясни", "объясни", "пожелай",
)

ANSWER_STARTERS = [
    r"^да\b", r"^нет\b", r"^хорошо\b", r"^ладно\b", r"^ок(ей)?\b", r"^понял\b", r"^согласен\b",
    r"^согласна\b", r"^верно\b", r"^именно\b", r"^конечно\b", r"^готово\b", r"^принято\b",
]

ANSWER_PATTERNS = [
    r"\bя\s+(сделаю|подготовлю|отправлю|проверю|исправлю|обновлю|настрою|доделаю|согласую)\b",
    r"\bмы\s+(сделаем|подготовим|отправим|проверим|исправим|обновим)\b",
    r"\b(сделано|готово|выполнено|завершено|загружено|отправлено)\b",
    r"\bза\s+это\s+отвечает\b",
    r"\bуже\s+(готов|готова|сделал|сделала|отправил|отправила)\b",
]

PROGRESS_PATTERNS = [
    r"\bесть\s+умерен(ный|ное|ная)\s+прогресс\b",
    r"\bмы\s+уже\s+синхронизировались\b",
    r"\bсейчас\s+коротко\s+расскажу\b",
    r"\bна\s+прошлой\s+неделе\s+мы\s+обсуждали\b",
    r"\bслово\s+передаю\b",
    r"\bобновление\s+по\s+своему\s+блоку\b",
    r"\bпрогресс\s+хороший\b",
]

SPEAKER_PREFIX_RE = re.compile(r"^\s*(Говорящий|Спикер|Speaker)\s*\d*\s*:\s*", re.I)
PERSON_PREFIX_RE = re.compile(r"^\s*([А-ЯЁ][а-яё]+)\s*,\s*(.+)$")
TASK_ACTION_VERBS = {
    "сделай", "подготовь", "проверь", "отправь", "создай", "добавь", "исправь",
    "обнови", "запусти", "напиши", "собери", "оформи", "согласуй", "перешли",
    "загрузи", "выгрузи", "протестируй", "закрой", "открой", "поставь",
    "назначь", "проведи", "созвонись", "уточни", "напомни", "доделай",
}

TASK_OBJECT_WORDS = {
    "отчет", "отчёт", "презентац", "документ", "таблиц", "файл", "договор",
    "задач", "таск", "тикет", "баг", "ошиб", "релиз", "деплой", "сервер",
    "письмо", "сообщение", "заявк", "макет", "дизайн", "код", "правк",
    "правки", "план", "созвон", "встреч", "аналитик", "дашборд", "экспорт",
    "импорт", "интеграц", "тест", "провер", "апдейт", "беклог", "бэклог",
    "api", "endpoint", "эндпоинт", "база", "данные", "клиент", "пользователь",
}

TASK_CONTEXT_WORDS = {
    "проект", "команда", "клиент", "работ", "прод", "продукт", "сервис",
    "спринт", "релиз", "созвон", "митинг", "встреч", "дедлайн", "срок",
    "ответствен", "исполнитель", "задач", "таск", "тикет", "беклог", "бэклог",
}

CASUAL_IMPERATIVE_PATTERNS = [
    r"\bонлайн-марафон\b",
    r"\bмарафон\s+желаний\b",
    r"\bпожелай\b",
    r"\bобнимите\b",
    r"\bрасскажи\s+словами\b",
    r"\bсмотри\b",
    r"\bсмотрите\b",
    r"\bпокажи\b",
    r"\bснимите\b",
    r"\bиди\b",
    r"\bидём\b",
    r"\bпойдём\b",
    r"\bдавай\s+другим\s+людям\s+понять\b",
]

QUESTION_ECHO_PATTERNS = [
    r"^о\s+ч[её]м\s+я\b",
    r"^что\s+я\b",
    r"^куда\s+бы\s+я\b",
    r"^как\s+бы\s+я\b",
    r"^с\s+кем\s+бы\s+я\b",
    r"^почему\s+я\b",
    r"^зачем\s+я\b",
]

QUESTION_BAD_EXACT = {
    "вы слышали",
    "вы слышали?",
    "да",
    "да?",
    "правда",
    "правда?",
    "может быть",
    "может быть?",
    "как красиво",
    "как красиво?",
    "патриоты",
    "патриоты?",
    "мечт",
    "мечт?",
    "мечтаний",
    "мечтаний?",
    "в сша, да",
    "в сша, да?",
    "почему я спрашиваю",
    "почему я спрашиваю?",
    "ну и почему он собственно используется",
    "ну и почему он собственно используется?",
    "ну и почему он используется",
    "ну и почему он используется?",
}

QUESTION_BAD_ENDINGS = [
    r",?\s+да\?$",
    r",?\s+правда\?$",
    r",?\s+наверное\?$",
    r",?\s+может\s+быть\?$",
    r"\bвы\s+слышали\?$",
]

MAIN_QUESTION_PATTERNS = [
    r"^следующий\s+вопрос\b.*\?",
    r"^если\s+бы\b",
    r"^расскажи\b",
    r"^поясни\b",
    r"^объясни\b",
    r"^пожелай\s+что-нибудь\b",
    r"^дай\s+совет\b",
    r"^самое\s+яркое\s+воспоминание\b",
    r"^самые\s+яркие\s+воспоминания\b",
    r"^с\s+какими\b",
    r"\bчто\s+для\s+тебя\b",
    r"\bчто\s+ты\s+думаешь\b",
    r"\bчто\s+ты\s+знаешь\b",
    r"\bчто\s+ты\s+имеешь\s+в\s+виду\b",
    r"\bкакое\s+у\s+тебя\s+мнение\b",
    r"\bна\s+что\s+бы\s+ты\b",
    r"\bгде\s+бы\s+ты\b",
    r"\bо\s+ч[её]м\s+ты\b",
    r"\bс\s+какими\s+\S+",
]

FOLLOWUP_QUESTION_EXACT = {
    "почему",
    "зачем",
    "как",
    "какой",
    "какая",
    "какие",
    "где именно",
    "что именно",
    "почему именно",
    "что это такое",
    "почему дурацкую",
    "почему невозможно",
    "каких проблем",
    "каких проблем ещё раз",
    "почему тебя это останавливает",
    "не хотят или хотят",
}

FOLLOWUP_QUESTION_PATTERNS = [
    r"^почему\s+\S+",
    r"^зачем\s+\S+",
    r"^кто\s+\S+",
    r"^сколько\s+\S+",
    r"^где\s+именно\b",
    r"^что\s+именно\b",
    r"^какой\s+\S+",
    r"^какая\s+\S+",
    r"^какие\s+\S+",
    r"^можешь\s+\S+",
]


def contains_any_stem(text: str, stems: set[str]) -> bool:
    lower = text.lower()
    return any(stem in lower for stem in stems)


def has_task_object(text: str) -> bool:
    return contains_any_stem(text, TASK_OBJECT_WORDS)


def has_task_context(text: str) -> bool:
    return contains_any_stem(text, TASK_CONTEXT_WORDS)


def has_deadline_phrase(text: str) -> bool:
    lower = text.lower()
    patterns = [
        r"\bдо\s+\d{1,2}[:.]\d{2}\b",
        r"\bдо\s+\d{1,2}[./]\d{1,2}(?:[./]\d{2,4})?\b",
        r"\bдо\s+(понедельника|вторника|среды|четверга|пятницы|субботы|воскресенья)\b",
        r"\bдо\s+(завтра|вечера|утра|обеда|релиза|конца\s+дня|конца\s+недели|следующей\s+встречи)\b",
        r"\bк\s+(утру|вечеру|обеду|созвону|завтрашнему\s+созвону|релизу|дедлайну)\b",
        r"\bсегодня\b",
        r"\bзавтра\b",
        r"\bна\s+этой\s+неделе\b",
        r"\bдо\s+конца\b",
    ]
    return any(re.search(p, lower) for p in patterns)


def has_responsible_signal(text: str) -> bool:
    lower = text.lower()
    return bool(
        PERSON_PREFIX_RE.match(text)
        or re.search(r"\bза\s+это\s+отвечает\b", lower)
        or re.search(r"\bответственн(ый|ая|ое|ые)\b", lower)
        or re.search(r"\bисполнитель\b", lower)
    )


def is_casual_imperative(text: str) -> bool:
    lower = strip_speaker_prefix(text).lower()
    return any(re.search(p, lower) for p in CASUAL_IMPERATIVE_PATTERNS)


def get_question_kind(text: str) -> str:
    raw = strip_speaker_prefix(text)
    lower = raw.lower().strip()
    clean = lower.strip(" .,!?:;—-")

    if not raw:
        return "not_question"

    if clean in QUESTION_BAD_EXACT:
        return "not_question"

    if any(re.search(p, lower) for p in QUESTION_BAD_ENDINGS):
        return "not_question"

    if is_filler_heavy(raw):
        return "not_question"

    if any(re.search(p, clean) for p in QUESTION_ECHO_PATTERNS):
        return "echo"

    if clean in FOLLOWUP_QUESTION_EXACT:
        return "followup"

    if any(re.search(p, clean) for p in MAIN_QUESTION_PATTERNS):
        return "main"

    if any(re.search(p, clean) for p in FOLLOWUP_QUESTION_PATTERNS):
        return "followup"

    if raw.endswith("?"):
        if re.search(
            r"^(почему|зачем|как|какой|какая|какие|где|когда|кто|что|сколько|можешь)\b",
            clean,
        ):
            return "followup"

    return "not_question"


def is_real_task_candidate(text: str, analysis: Optional[dict] = None) -> bool:
    clean = clean_task_text(text) if "clean_task_text" in globals() else strip_speaker_prefix(text)
    lower = clean.lower().strip(" .,!?:;—-")

    if not clean:
        return False

    if is_casual_imperative(clean):
        return False

    if is_filler_heavy(clean):
        return False

    first_word = lower.split()[0] if lower.split() else ""

    strong_context = (
        has_task_object(clean)
        or has_deadline_phrase(clean)
        or has_responsible_signal(clean)
        or has_task_context(clean)
    )

    direct_imperative = first_word in TASK_ACTION_VERBS and strong_context

    necessity = bool(re.search(
        r"^(надо|нужно|нужна|нужен|нужны|важно|стоит)\s+"
        r"(сделать|подготовить|проверить|отправить|создать|добавить|исправить|"
        r"обновить|запустить|написать|собрать|оформить|согласовать|загрузить|"
        r"выгрузить|протестировать|закрыть|открыть|назначить|провести|уточнить|доделать)\b",
        lower,
    ))

    collective = bool(re.search(
        r"^(давайте|давай)\s+"
        r"(сделаем|подготовим|проверим|отправим|создадим|добавим|исправим|"
        r"обновим|запустим|напишем|соберём|оформим|согласуем|загрузим|"
        r"протестируем|закроем|откроем|уточним|доделаем)\b",
        lower,
    ))

    if direct_imperative or necessity or collective:
        return True

    if analysis and analysis.get("is_task"):
        reasons = set(analysis.get("reasons") or [])
        if strong_context and (
            "deadline" in reasons
            or "responsible" in reasons
            or "imperative" in reasons
            or float(analysis.get("score", 0) or 0) >= 0.82
        ):
            return True

    return False
MIN_QA_CHARS = 18


def normalize_text(text: str) -> str:
    text = str(text or "").replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text.strip(" -–—")


SENTENCE_SPLIT_RE = re.compile(r"(?<=[.!?…])\s+|\s*[;]\s+")


def split_text_units(text: str, max_words: int = MAX_SEGMENT_WORDS) -> list[str]:
    """
    Дробит длинный текст Whisper-сегмента на более короткие смысловые части.
    Это критично для выделения задач: в длинном сегменте задача часто теряется
    среди вводных слов и соседних реплик.
    """
    text = normalize_text(text)

    if not text:
        return []

    raw_parts = [p.strip(" ,;:-–—") for p in SENTENCE_SPLIT_RE.split(text) if p.strip(" ,;:-–—")]

    # Если пунктуации почти нет, дополнительно режем по запятым и союзам-маркерам.
    if len(raw_parts) <= 1 and len(text.split()) > max_words:
        raw_parts = [p.strip(" ,;:-–—") for p in re.split(r"\s*,\s+|\s+и\s+потом\s+|\s+потом\s+|\s+далее\s+", text) if p.strip(" ,;:-–—")]

    parts: list[str] = []

    for part in raw_parts or [text]:
        words = part.split()

        if len(words) <= max_words:
            if len(part) >= MIN_SEGMENT_CHARS:
                parts.append(part)
            continue

        # Режем слишком длинные куски блоками без overlap, чтобы не плодить дубли задач.
        for i in range(0, len(words), max_words):
            chunk = " ".join(words[i:i + max_words]).strip(" ,;:-–—")
            if len(chunk) >= MIN_SEGMENT_CHARS:
                parts.append(chunk)

    return parts or ([text] if len(text) >= MIN_SEGMENT_CHARS else [])


def split_whisper_segment(start: float, end: float, text: str) -> list[dict]:
    """
    Превращает один Whisper-сегмент в несколько коротких сегментов.
    Таймкоды распределяются приблизительно пропорционально длине текста.
    """
    text = normalize_text(text)

    if not SPLIT_TRANSCRIPT_SEGMENTS:
        return [{"start": round(start, 2), "end": round(end, 2), "text": text}] if text else []

    units = split_text_units(text)

    if len(units) <= 1:
        return [{"start": round(start, 2), "end": round(end, 2), "text": text}] if text else []

    duration = max(float(end) - float(start), 0.01)
    total_chars = max(sum(len(u) for u in units), 1)
    cursor = float(start)
    out: list[dict] = []

    for idx, unit in enumerate(units):
        if idx == len(units) - 1:
            unit_end = float(end)
        else:
            unit_duration = duration * (len(unit) / total_chars)
            unit_end = min(float(end), cursor + max(unit_duration, 0.15))

        out.append({
            "start": round(cursor, 2),
            "end": round(unit_end, 2),
            "text": unit,
        })
        cursor = unit_end

    return out


TASK_SOFT_ACTION_WORDS = {
    "сделать", "подготовить", "проверить", "отправить", "создать", "добавить", "исправить",
    "обновить", "запустить", "написать", "собрать", "оформить", "согласовать", "загрузить",
    "выгрузить", "протестировать", "закрыть", "открыть", "назначить", "провести", "уточнить",
    "доделать", "доработать", "разработать", "реализовать", "настроить", "посмотреть",
    "переписать", "сформировать", "описать", "подобрать", "разместить", "получить",
    "запросить", "забронировать", "сверить", "переслать", "созвониться", "обсудить",
    "передать", "зафиксировать", "прикрепить", "вынести", "перерисовать", "заменить",
}

PROJECT_TASK_HINTS = {
    "презентац", "отч", "документ", "таблиц", "слайд", "вкр", "код", "модель", "датасет",
    "база", "интерфейс", "фронт", "бэк", "api", "эндпоинт", "сервер", "деплой", "ошиб",
    "правк", "задач", "срок", "дедлайн", "файл", "проект", "раздел", "тест", "метрик",
    "скрин", "архитектур", "схем", "диаграм", "экспорт", "пользователь", "авторизац",
}


def has_soft_action(text: str) -> bool:
    lower = text.lower()
    if any(re.search(rf"\b{re.escape(v)}\b", lower) for v in TASK_SOFT_ACTION_WORDS):
        return True
    return any(re.search(rf"\b{re.escape(v)}\b", lower) for v in TASK_ACTION_VERBS)


def has_project_task_hint(text: str) -> bool:
    lower = text.lower()
    return any(h in lower for h in PROJECT_TASK_HINTS) or has_task_object(text) or has_task_context(text)


def looks_like_task_soft(text: str) -> bool:
    """
    Мягкий фильтр задач для реальных встреч.
    Он не должен быть единственным источником истины, но помогает принять
    уверенное предсказание модели и не пропустить разговорные формулировки.
    """
    clean = strip_speaker_prefix(text)
    lower = clean.lower().strip(" .,!?:;—-")

    if not clean or len(clean.split()) < 2:
        return False

    if is_filler_heavy(clean) or is_casual_imperative(clean):
        return False

    if get_question_kind(clean) in {"main", "followup", "echo"}:
        return False

    if clean.endswith("?"):
        return False

    if re.search(r"\b(надо|нужно|нужна|нужен|нужны|необходимо|требуется|следует|важно|стоит)\b", lower) and has_soft_action(clean):
        return True

    if re.search(r"\b(давай|давайте|прошу|планируем|планируется|надо бы|нужно будет)\b", lower) and has_soft_action(clean):
        return True

    first = lower.split()[0] if lower.split() else ""
    if (first in TASK_ACTION_VERBS or first in TASK_SOFT_ACTION_WORDS) and (has_project_task_hint(clean) or has_deadline_phrase(clean)):
        return True

    if has_soft_action(clean) and (has_deadline_phrase(clean) or has_responsible_signal(clean) or has_project_task_hint(clean)):
        return True

    return False


def strip_speaker_prefix(text: str) -> str:
    text = SPEAKER_PREFIX_RE.sub("", text)
    return normalize_text(text)


def is_filler_heavy(text: str) -> bool:
    words = [
        w.strip(".,!?;:()[]\"'«»").lower()
        for w in text.split()
        if w.strip(".,!?;:()[]\"'«»")
    ]
    if len(words) <= 3:
        return False
    filler_count = sum(1 for w in words if w in FILLER_WORDS)
    return filler_count / max(len(words), 1) >= 0.45


def is_progress_text(text: str) -> bool:
    lower = text.lower()
    return any(re.search(p, lower) for p in PROGRESS_PATTERNS)


def extract_deadline(text: str) -> Optional[str]:
    analysis = analyze_task_candidate(text)
    reasons = analysis.get("reasons", [])
    for r in reasons:
        if r == "deadline":
            lower = text.lower()
            patterns = [
                r"\bдо\s+\d{1,2}[:.]\d{2}\b",
                r"\bдо\s+\d{1,2}[./]\d{1,2}(?:[./]\d{2,4})?\b",
                r"\bдо\s+(понедельника|вторника|среды|четверга|пятницы|субботы|воскресенья)\b",
                r"\bдо\s+(завтра|вечера|утра|обеда|релиза|конца\s+дня|конца\s+недели|следующей\s+встречи)\b",
                r"\bк\s+(утру|вечеру|обеду|созвону|завтрашнему\s+созвону)\b",
                r"\bсегодня\b",
                r"\bзавтра\b",
                r"\bна\s+этой\s+неделе\b",
                r"\bпосле\s+обеда\b",
            ]
            for p in patterns:
                m = re.search(p, lower)
                if m:
                    return m.group(0)
    return None


def extract_responsible(text: str) -> Optional[str]:
    m = PERSON_PREFIX_RE.match(text)
    if m:
        return m.group(1)

    if NATASHA_AVAILABLE and ner_tagger and segmenter:
        try:
            doc = Doc(text)
            doc.segment(segmenter)
            doc.tag_ner(ner_tagger)
            for span in doc.spans:
                if span.type == "PER":
                    span.normalize(morph_vocab)
                    return span.normal if getattr(span, "normal", None) else span.text
        except Exception:
            pass

    m = re.search(r"\bза\s+это\s+отвечает\s+([А-ЯЁ][а-яё]+)\b", text)
    if m:
        return m.group(1)

    return None


def classify_by_rules(text: str) -> tuple[str, float, dict]:
    text = strip_speaker_prefix(text)

    debug = {
        "filler": is_filler_heavy(text),
        "progress": is_progress_text(text),
    }

    task_analysis = get_task_analysis_safe(text)
    debug["task_rule_engine"] = task_analysis

    if task_analysis.get("is_task"):
        return "task", max(0.91, float(task_analysis.get("score", 0.91))), debug

    question_kind = get_question_kind(text)
    debug["question_kind"] = question_kind

    if question_kind in {"main", "followup"}:
        return "question", 0.97, debug

    if question_kind in {"echo", "not_question"} and len(text) < MIN_QA_CHARS:
        if looks_like_short_answer(text):
            return "answer", 0.90, debug
        return "other", 0.95, debug

    if is_filler_heavy(text):
        return "other", 0.95, debug

    lower = text.lower()

    if any(re.search(p, lower) for p in ANSWER_STARTERS) or any(re.search(p, lower) for p in ANSWER_PATTERNS):
        if is_good_answer(text):
            return "answer", 0.93, debug

    if looks_like_short_answer(text):
        return "answer", 0.90, debug

    if is_progress_text(text):
        return "other", 0.97, debug

    return "other", 0.9, debug

def batch_classifier_predict(texts: list[str]) -> list[tuple[str, float]]:
    if not (TORCH_AVAILABLE and classifier_tokenizer and classifier_model):
        return [("other", 0.0) for _ in texts]

    device = next(classifier_model.parameters()).device
    out: list[tuple[str, float]] = []

    for i in range(0, len(texts), INFERENCE_BATCH_SIZE):
        batch = texts[i:i + INFERENCE_BATCH_SIZE]
        enc = classifier_tokenizer(
            batch,
            truncation=True,
            padding=True,
            max_length=128,
            return_tensors="pt",
        )
        enc = {k: v.to(device) for k, v in enc.items()}

        with torch.no_grad():
            logits = classifier_model(**enc).logits
            probs = F.softmax(logits, dim=-1)
            confs, idxs = probs.max(dim=-1)

        for conf, idx in zip(confs.tolist(), idxs.tolist()):
            out.append((LABELS.get(idx, "other"), float(conf)))

    return out

TASK_MODEL_MIN_CONF = float(os.getenv("CLASSIFIER_TASK_CONFIDENCE_THRESHOLD", "0.82"))


def has_actionable_task_signal(text: str) -> bool:
    clean = strip_speaker_prefix(str(text or ""))
    lower = clean.lower().replace("ё", "е").strip(" .,!?:;—-")

    if not lower:
        return False

    if is_filler_heavy(clean):
        return False

    if is_casual_imperative(clean):
        return False

    if get_question_kind(clean) in {"main", "followup", "echo"}:
        return False

    # Отсекаем рассуждения, гипотезы и технические объяснения без поручения
    reject_patterns = [
        r"\bзадача\s+(эксперта|интерпретации|оптимизации|классификации|регрессии)\b",
        r"\bэто\s+(ваша|наша)\s+задача\b",
        r"\bможем\s+(посчитать|использовать|построить|попробовать)\b",
        r"\bмы\s+можем\s+(посчитать|использовать|построить|попробовать)\b",
        r"\bя\s+(могу|не\s+смогу|думаю|полагаю|считаю|пробовал)\b",
        r"\bможно\s+(будет\s+)?(посмотреть|попробовать|использовать|сказать)\b",
        r"\bполучается\b",
        r"\bв\s+принципе\b",
        r"\bкак\s+гипотеза\b",
        r"\bэто\s+как\s+гипотеза\b",
        r"\bсобственно\b",
        r"\bпроксимировать\b",
        r"\bинтерполировать\b",
        r"\bгенерализовать\b",
        r"\bне\s+можем\s+использовать\b",
        r"\bне\s+смогу\b",
    ]

    if any(re.search(p, lower) for p in reject_patterns):
        # Исключение: если прямо сказано "нужно/надо/требуется проверить/сделать"
        if not re.search(r"\b(нужно|надо|необходимо|требуется|следует)\b", lower):
            return False

    # Явные формулировки задач
    strong_patterns = [
        r"^(подготовить|проверить|исправить|добавить|разработать|реализовать|"
        r"сделать|согласовать|оформить|протестировать|посмотреть|уточнить|"
        r"сформировать|описать|собрать|отправить|разобраться|попробовать|"
        r"провести|выделить|разбить|сопоставить|найти|поискать|доработать)\b",

        r"\b(нужно|надо|необходимо|требуется|следует|важно)\s+"
        r"(подготовить|проверить|исправить|добавить|разработать|реализовать|"
        r"сделать|согласовать|оформить|протестировать|посмотреть|уточнить|"
        r"сформировать|описать|собрать|отправить|разобраться|попробовать|"
        r"провести|выделить|разбить|сопоставить|найти|поискать|доработать)\b",

        r"\b(давайте|давай)\s+"
        r"(подготовим|проверим|исправим|добавим|разработаем|реализуем|"
        r"сделаем|согласуем|оформим|протестируем|посмотрим|уточним|"
        r"сформируем|опишем|соберем|соберём|разберемся|разберёмся|"
        r"попробуем|проведем|проведём|выделим|разобьем|разобьём|сопоставим)\b",

        r"\bк\s+следующей\s+встрече\b",
        r"\bдо\s+(пятницы|четверга|среды|понедельника|завтра|конца\s+недели|следующей\s+встречи)\b",
        r"\bна\s+следующей\s+неделе\b",
    ]

    if any(re.search(p, lower) for p in strong_patterns):
        return True

    # Дополнительный мягкий случай: есть действие + объект проекта
    action_words = [
    "проверить", "разобраться", "попробовать", "разбить",
    "выделить", "сопоставить", "уточнить", "подготовить", "доработать",
    "исправить", "найти", "поискать", "провести",
    ]

    object_words = [
        "модель", "параметр", "параметры", "кейсы", "данные", "формулы",
        "расхождение", "устойчивость", "точность", "группы", "скин",
        "трещин", "аппроксимацию", "коэффициенты", "вкр", "презентацию",
        "документ", "отчет", "отчёт",
    ]

    has_action = any(re.search(rf"\b{re.escape(w)}\b", lower) for w in action_words)
    has_object = any(re.search(rf"\b{re.escape(w)}\b", lower) for w in object_words)

    return has_action and has_object

def final_classify(texts: list[str]) -> list[dict]:
    rule_results = [classify_by_rules(t) for t in texts]
    clf_results = batch_classifier_predict(texts)
    final: list[dict] = []

    for text, (rule_label, rule_conf, debug), (clf_label, clf_conf) in zip(texts, rule_results, clf_results):
        clean = strip_speaker_prefix(text)
        task_analysis = get_task_analysis_safe(clean)
        question_kind = get_question_kind(clean)

        label = "other"
        confidence = 0.0
        source = "fallback_other"

        actionable_task = has_actionable_task_signal(clean)

        # 1. Вопросы
        if (
    question_kind in {"main", "followup"}
            or is_good_question(clean)
            or (
                clean.strip().endswith("?")
                and re.search(
                    r"^\s*(почему|зачем|как|какой|какая|какое|какие|где|когда|кто|что|сколько|можно ли|нужно ли|а какие|а как|а почему|а что|ну и почему|в чем|в чём)\b",
                    clean.lower(),
                )
            )
            or (
                clf_label == "question"
                and clf_conf >= 0.90
                and is_good_question(clean)
            )
        ):
            label = "question"
            confidence = max(float(rule_conf), float(clf_conf), 0.91)
            source = "question_detected"

        # 2. Ответы
        elif rule_label == "answer" and is_good_answer(clean):
            label = "answer"
            confidence = float(rule_conf)
            source = "rules_answer"

        elif clf_label == "answer" and clf_conf >= CLASSIFIER_ANSWER_CONFIDENCE and is_good_answer(clean):
            label = "answer"
            confidence = float(clf_conf)
            source = "classifier_answer"

        # 3. Задачи
        elif task_analysis.get("is_task") and actionable_task:
            label = "task"
            confidence = max(float(rule_conf), float(task_analysis.get("score", 0.91)))
            source = "rules_task"

        elif clf_label == "task" and clf_conf >= TASK_MODEL_MIN_CONF and actionable_task:
            label = "task"
            confidence = float(clf_conf)
            source = "classifier_task_actionable"

        # 4. Остальное
        elif clf_label == "other" and clf_conf >= CLASSIFIER_CONFIDENCE:
            label = "other"
            confidence = float(clf_conf)
            source = "classifier_other"

        else:
            label = "other"
            confidence = max(float(clf_conf), float(rule_conf), 0.49)
            source = "safe_other"

        # Защита от ложных задач
        if label == "task":
            if not actionable_task:
                label = "other"
                confidence = 0.49
                source = "task_actionable_reject"

            if is_good_question(clean) or question_kind in {"main", "followup", "echo"} or clean.strip().endswith("?"):
                label = "other"
                confidence = 0.49
                source = "task_question_reject"

            if is_filler_heavy(clean):
                label = "other"
                confidence = 0.49
                source = "task_filler_reject"

        final.append({
            "text": clean,
            "label": label,
            "confidence": round(float(confidence), 4),
            "source": source,
            "debug": {
                **debug,
                "classifier_label": clf_label,
                "classifier_confidence": round(float(clf_conf), 4),
                "rule_label": rule_label,
                "rule_confidence": round(float(rule_conf), 4),
                "actionable_task": actionable_task,
            },
            "task_debug": task_analysis,
        })

    return final


def batch_sentiment(texts: list[str]) -> list[dict]:
    if not sentiment_pipe:
        return [{"label": "neutral", "score": 0.0} for _ in texts]

    results: list[dict] = []

    for i in range(0, len(texts), INFERENCE_BATCH_SIZE):
        batch = texts[i:i + INFERENCE_BATCH_SIZE]

        try:
            preds = sentiment_pipe(batch)
        except Exception:
            preds = [{"label": "neutral", "score": 0.0} for _ in batch]

        for pred in preds:
            label = str(pred.get("label", "neutral")).lower()
            score = float(pred.get("score", 0.0))

            if "neg" in label:
                results.append({"label": "negative", "score": round(-score, 4)})
            elif "pos" in label:
                results.append({"label": "positive", "score": round(score, 4)})
            else:
                results.append({"label": "neutral", "score": 0.0})

    return results


def is_good_question(text: str) -> bool:
    text = strip_speaker_prefix(text)

    if not text:
        return False

    lower = text.lower().strip()
    clean = lower.strip(" .,!?:;—-")

    if is_filler_heavy(text):
        return False

    bad_questions = {
        "вы слышали",
        "вы слышали?",
        "да",
        "да?",
        "правда",
        "правда?",
        "может быть",
        "может быть?",
        "как красиво",
        "как красиво?",
        "патриоты",
        "патриоты?",
        "мечт",
        "мечт?",
        "мечтаний",
        "мечтаний?",
        "коммунных",
        "коммунных?",
        "в сша, да",
        "в сша, да?",
        "на такие простые, как бы, работы, да",
        "на такие простые, как бы, работы, да?",
    }

    if clean in bad_questions:
        return False

    bad_endings = [
        "вы слышали?",
        ", да?",
        " да?",
        "наверное?",
        "может быть?",
        "будет марафон?",
    ]

    if any(lower.endswith(x) for x in bad_endings):
        return False

    if looks_like_short_question(text):
        return True

    if len(text) < QA_MIN_QUESTION_CHARS:
        return False

    leading = r"^(?:а\s+|но\s+|и\s+|ну\s+|вот\s+|во-первых,\s+|во-первых\s+)?"

    strong_question_patterns = [
        leading + r"(почему|зачем|кто|сколько|где|когда|как|какой|какая|какое|какие)\b",
        leading + r"(можешь|можете|можем|можно\s+ли|нужно\s+ли|успеем\s+ли)\b",
        leading + r"(расскажи|поясни|объясни|пожелай)\b",
        r"^если\s+бы\b",
        r"\bчто\s+ты\s+думаешь\b",
        r"\bчто\s+ты\s+знаешь\b",
        r"\bчто\s+для\s+тебя\b",
        r"\bчто\s+ты\s+имеешь\s+в\s+виду\b",
        r"\bкакое\s+у\s+тебя\s+мнение\b",
        r"\bс\s+какими\b",
        r"\bна\s+что\s+бы\s+ты\b",
        r"\bгде\s+бы\s+ты\b",
        r"\bо\s+ч[её]м\s+ты\b",
        r"\bсамое\s+яркое\s+воспоминание\b",
        r"\bсамые\s+яркие\s+воспоминания\b",
        r"\bдай\s+совет\b",
    ]

    if any(re.search(p, lower) for p in strong_question_patterns):
        return True

    if text.endswith("?"):
        weak_question_patterns = [
            r"\bили\b",
            r"\bчто\b",
            r"\bкак\b",
            r"\bгде\b",
            r"\bпочему\b",
            r"\bзачем\b",
            r"\bможем\b",
            r"\bможешь\b",
            r"\bможете\b",
        ]
        return any(re.search(p, lower) for p in weak_question_patterns)

    no_mark_question_patterns = [
        r"^самые\s+яркие\s+воспоминания\b",
        r"^самое\s+яркое\s+воспоминание\b",
        r"^любимое\s+воспоминание\b",
        r"^сложный\s+период\b",
        r"^самый\s+сложный\s+период\b",
    ]

    return any(re.search(p, lower) for p in no_mark_question_patterns)


def is_good_answer(text: str) -> bool:
    text = strip_speaker_prefix(text)

    if not text:
        return False

    lower = text.lower().strip()
    clean = lower.strip(" .,!?:;—-")

    if is_filler_heavy(text):
        return False

    if is_good_question(text):
        return False

    if text.endswith("?"):
        return False

    bad_short_answers = {
        "да",
        "нет",
        "ага",
        "угу",
        "понятно",
        "хорошо",
        "ладно",
        "ок",
        "окей",
        "возможно",
        "не знаю",
        "так",
        "вот",
        "супер",
        "правда",
    }

    if clean in bad_short_answers:
        return False

    bad_question_like = [
        r"\bчто\s+ты\b",
        r"\bчто\s+вы\b",
        r"\bкакое\s+у\s+тебя\s+мнение\b",
        r"\bкакое\s+у\s+вас\s+мнение\b",
        r"\bпочему\b",
        r"\bзачем\b",
        r"\bкак\s+попасть\b",
        r"\bне\s+хотят\s+или\s+хотят\b",
        r"\bможет\s+быть\b",
        r"\bда\?$",
    ]

    if any(re.search(p, lower) for p in bad_question_like):
        return False

    if looks_like_short_answer(text):
        return True

    if len(text) < QA_MIN_ANSWER_CHARS:
        return False

    if any(re.search(p, clean) for p in ANSWER_STARTERS):
        return True

    if any(re.search(p, clean) for p in ANSWER_PATTERNS):
        return True

    if is_progress_text(text):
        return False

    answer_markers = [
        r"\bя\s+(думаю|считаю|сказала|сказал|жила|пошла|хочу|знаю|понимаю|мечтаю)\b",
        r"\bмне\s+(кажется|важно|комфортно|нравится|не\s+нравится)\b",
        r"\bдля\s+меня\b",
        r"\bэто\s+(было|есть|школа|тихий|важно|сложно|здорово)\b",
        r"\bпотому\s+что\b",
        r"\bна\s+самом\s+деле\b",
        r"\bнапример\b",
        r"\bну\s+да\b",
        r"\bда,\s+я\b",
    ]

    if any(re.search(p, lower) for p in answer_markers):
        return True

    return len(text.split()) >= 6

def is_imperative_task(text: str) -> bool:
    return is_real_task_candidate(text)

def is_imperative_task_raw(text: str) -> bool:
    clean = strip_speaker_prefix(text)
    lower = clean.lower().strip(" .,!?:;—-")
    first_word = lower.split()[0] if lower.split() else ""

    if first_word in TASK_ACTION_VERBS:
        return True

    patterns = [
        r"^(надо|нужно|нужна|нужен|нужны|важно|стоит)\s+\S+",
        r"^(давайте|давай)\s+\S+",
    ]

    return any(re.search(p, lower) for p in patterns)
def get_task_analysis_safe(text: str) -> dict:
    clean = clean_task_text(text) if "clean_task_text" in globals() else strip_speaker_prefix(text)

    try:
        analysis = analyze_task_candidate(clean)
    except Exception:
        analysis = {
            "is_task": False,
            "score": 0.0,
            "reasons": [],
        }

    if is_real_task_candidate(clean, analysis):
        reasons = list(analysis.get("reasons") or [])

        if not reasons:
            reasons = ["global_task_filter"]

        if is_imperative_task_raw(clean) and "imperative" not in reasons:
            reasons.append("imperative")

        return {
            **analysis,
            "is_task": True,
            "score": max(float(analysis.get("score", 0.0) or 0.0), 0.91),
            "reasons": reasons,
        }

    return {
        **analysis,
        "is_task": False,
        "score": min(float(analysis.get("score", 0.0) or 0.0), 0.49),
    }


def looks_like_short_question(text: str) -> bool:
    text = strip_speaker_prefix(text)
    lower = text.lower().strip(" .,!?:;—-")

    short_questions = {
        "почему",
        "зачем",
        "как",
        "какой",
        "какая",
        "какое",
        "какие",
        "где",
        "когда",
        "кто",
        "что",
        "где именно",
        "что именно",
        "почему именно",
        "что это такое",
        "что ты имеешь в виду",
        "не хотят или хотят",
        "почему невозможно",
        "каких проблем",
        "каких проблем ещё раз",
        "каких проблем еще раз",
        "почему дурацкую",
        "почему тебя это останавливает",
        "где именно",
        "как попасть в cern",
        "о чём я жалею",
        "о чем я жалею",
        "о чём я мечтаю",
        "о чем я мечтаю",
    }

    bad_short_questions = {
        "как красиво",
        "патриоты",
        "мечт",
        "мечтаний",
        "вы слышали",
        "да",
        "правда",
        "коммунных",
        "в сша, да",
    }

    if lower in bad_short_questions:
        return False

    if lower in short_questions:
        return True

    if text.endswith("?") and len(lower.split()) <= 8:
        if lower.endswith("да") or lower.endswith("наверное") or lower.endswith("марафон"):
            return False
        if re.search(r"\b(почему|зачем|как|какой|какая|какое|какие|где|когда|кто|что|можем|можешь|можете|можно)\b", lower):
            return True

    return False


def looks_like_short_answer(text: str) -> bool:
    text = strip_speaker_prefix(text)
    lower = text.lower().strip(" .,!?:;—-")

    bad_too_short = {
        "да",
        "нет",
        "ага",
        "угу",
        "ок",
        "окей",
        "понятно",
        "хорошо",
        "ладно",
    }

    if lower in bad_too_short:
        return False

    short_answers = {
        "это тихий",
        "тихий",
        "медведей нет",
        "хотят",
        "не хотят",
        "да нет",
        "ну да",
        "это школа",
        "по математике нет",
        "это было сложно",
    }

    if lower in short_answers:
        return True

    if len(lower.split()) <= 5 and any(
        w in lower for w in ["тихий", "школа", "медведей", "хотят", "сложно"]
    ):
        return True

    return False


def is_topic_shift(text: str, answer_start: str = "") -> bool:
    lower = strip_speaker_prefix(text).lower().strip(" .,!?:;—-")

    topic_starters = [
        r"^следующий вопрос\b",
        r"^самые яркие воспоминания\b",
        r"^самое яркое воспоминание\b",
        r"^расскажи\b",
        r"^о ч[её]м ты\b",
        r"^с какими\b",
        r"^что ты думаешь\b",
        r"^что ты знаешь\b",
        r"^пожелай\b",
        r"^дай совет\b",
        r"^давай\b",
        r"^сейчас покажу\b",
        r"^подходим\b",
        r"^так смотрим\b",
        r"^теперь\b",
        r"^следующая тема\b",
        r"^если бы\b",
        r"^лена,\s+что\b",
    ]

    if any(re.search(p, lower) for p in topic_starters):
        return True

    if lower in {"так", "вот", "понятно", "супер", "хорошо", "следующий вопрос"}:
        return True

    return False
def normalized_for_match(text: str) -> str:
    text = strip_speaker_prefix(text).lower()
    text = re.sub(r"[^\w\sа-яё]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def is_echo_question(current: str, previous: str) -> bool:
    cur = normalized_for_match(current)
    prev = normalized_for_match(previous)

    if not cur or not prev:
        return False

    echo_patterns = [
        r"^о ч[её]м я\b",
        r"^что я\b",
        r"^почему я\b",
        r"^как я\b",
        r"^где я\b",
        r"^какое у меня\b",
    ]

    if any(re.search(p, cur) for p in echo_patterns):
        return True

    cur_words = set(cur.split())
    prev_words = set(prev.split())

    if not cur_words or not prev_words:
        return False

    overlap = len(cur_words & prev_words) / max(len(cur_words), 1)

    return overlap >= 0.65 and len(cur_words) <= len(prev_words) + 2


def is_followup_question(current: str, question_segments: list[dict]) -> bool:
    cur = strip_speaker_prefix(current).lower().strip(" .,!?:;—-")

    if not question_segments:
        return False

    first = strip_speaker_prefix(question_segments[0]["text"]).lower()

    followup_patterns = [
        r"^вот\s+что\s+ты\b",
        r"^какое\s+у\s+тебя\s+мнение\b",
        r"^что\s+ты\s+о\s+них\s+знаешь\b",
        r"^что\s+ты\s+думаешь\b",
        r"^во-первых\b",
    ]

    if any(re.search(p, cur) for p in followup_patterns):
        return True

    if "нынешних россиян" in first or "русских" in first or "россиянах" in first:
        if re.search(r"\b(что|какое|мнение|знаешь|думаешь)\b", cur):
            return True

    return False


def normalize_question_text(question_segments: list[dict]) -> str:
    parts = []

    for seg in question_segments:
        text = strip_speaker_prefix(seg["text"])
        low = text.lower().strip(" .,!?:;—-")

        if re.search(r"^о ч[её]м я\b", low):
            continue

        if parts and is_echo_question(text, parts[-1]):
            continue

        parts.append(text)

    if not parts and question_segments:
        parts.append(strip_speaker_prefix(question_segments[0]["text"]))

    text = " ".join(parts)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def dedupe_qa_pairs(pairs: list[dict]) -> list[dict]:
    result = []
    seen = set()

    for pair in pairs:
        q_raw = pair.get("question", "")
        a_raw = pair.get("answer", "")

        q = normalized_for_match(q_raw)
        a = normalized_for_match(a_raw)

        if not q or not a:
            continue

        if q == a:
            continue

        if is_good_question(a_raw):
            continue

        key = (q[:160], a[:180])

        if key in seen:
            continue

        duplicate = False

        for existing in result:
            eq = normalized_for_match(existing.get("question", ""))
            ea = normalized_for_match(existing.get("answer", ""))

            q_words = set(q.split())
            eq_words = set(eq.split())

            if q_words and eq_words:
                q_overlap = len(q_words & eq_words) / max(min(len(q_words), len(eq_words)), 1)
            else:
                q_overlap = 0

            if q_overlap >= 0.8 and a[:120] == ea[:120]:
                duplicate = True
                break

        if duplicate:
            continue

        seen.add(key)
        result.append(pair)

    return result

def dedupe_tasks(tasks: list[dict]) -> list[dict]:
    result: list[dict] = []

    for task in tasks:
        text = task.get("text", "")
        norm = normalize_task_text(text)

        if not norm:
            continue

        duplicate = False

        for existing in result:
            if text_similarity(text, existing.get("text", "")) >= 0.82:
                duplicate = True
                break

        if duplicate:
            continue

        result.append(task)

    return result


def normalize_task_text(text: str) -> str:
    text = strip_speaker_prefix(str(text or "")).lower()

    # Убираем имена/обращения в начале: "Алексей, Анна, подготовь..."
    text = re.sub(
        r"^(?:[а-яё]+,\s*){1,4}",
        "",
        text,
        flags=re.I,
    )

    # Приводим разные формы ё/е
    text = text.replace("ё", "е")

    # Убираем пунктуацию
    text = re.sub(r"[^\w\sа-яё]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    # Нормализуем частые формы глаголов действия
    replacements = {
        "подготовь": "подготовить",
        "подготовьте": "подготовить",
        "подготовим": "подготовить",
        "подготовить": "подготовить",
        "исправь": "исправить",
        "исправьте": "исправить",
        "исправим": "исправить",
        "исправить": "исправить",
        "проверь": "проверить",
        "проверьте": "проверить",
        "проверим": "проверить",
        "проверить": "проверить",
        "отправь": "отправить",
        "отправьте": "отправить",
        "отправим": "отправить",
        "отправить": "отправить",
        "добавь": "добавить",
        "добавьте": "добавить",
        "добавим": "добавить",
        "добавить": "добавить",
        "обнови": "обновить",
        "обновите": "обновить",
        "обновим": "обновить",
        "обновить": "обновить",
        "собери": "собрать",
        "соберите": "собрать",
        "соберем": "собрать",
        "соберём": "собрать",
        "собрать": "собрать",
    }

    stop_words = {
        "нужно", "надо", "важно", "стоит", "давайте", "давай",
        "пожалуйста", "тогда", "еще", "ещё", "ну", "вот", "там", "тут",
        "это", "этот", "эта", "эти", "типа", "как", "бы", "просто",
        "к", "до", "на", "по", "и", "или", "в", "во", "с", "со", "для",
        "будет", "будем", "можно", "уже",
        "алексей", "анна", "иван", "мария",
    }

    words = []
    for w in text.split():
        w = replacements.get(w, w)

        if w in stop_words:
            continue

        if len(w) <= 2:
            continue

        words.append(w)

    return " ".join(words)


def text_similarity(a: str, b: str) -> float:
    a_norm = normalize_task_text(a)
    b_norm = normalize_task_text(b)

    if not a_norm or not b_norm:
        return 0.0

    a_words = set(a_norm.split())
    b_words = set(b_norm.split())

    if not a_words or not b_words:
        return 0.0

    intersection = len(a_words & b_words)
    union = len(a_words | b_words)

    jaccard = intersection / union if union else 0.0
    containment = intersection / min(len(a_words), len(b_words))

    # Берем максимум: так одинаковая задача с лишним обращением/сроком не станет новой
    score = max(jaccard, containment)

    return round(score, 4)


def compare_tasks_with_previous(current_tasks: list[dict], previous_tasks: list[dict]) -> dict:
    repeated_tasks = []
    new_tasks = []
    potentially_closed_tasks = []

    matched_previous_indexes = set()

    # Для одинаковой повторной загрузки порог должен быть достаточно мягким
    SIMILARITY_THRESHOLD = 0.42

    for current_task in current_tasks:
        current_text = current_task.get("text", "")
        current_norm = normalize_task_text(current_text)

        best_match = None
        best_score = 0.0
        best_index = None

        for idx, previous_task in enumerate(previous_tasks):
            previous_text = previous_task.get("text", "")
            previous_norm = normalize_task_text(previous_text)

            if not current_norm or not previous_norm:
                continue

            score = text_similarity(current_text, previous_text)

            if score > best_score:
                best_score = score
                best_match = previous_task
                best_index = idx

        if best_match and best_score >= SIMILARITY_THRESHOLD:
            repeated_item = {
                "current_task": current_task,
                "previous_task": best_match,
                "similarity": best_score,
            }

            repeated_tasks.append(repeated_item)

            if best_index is not None:
                matched_previous_indexes.add(best_index)

            # Важно: помечаем статус прямо в текущей задаче
            current_task["status"] = "repeated"
            current_task["similarity"] = best_score
        else:
            current_task["status"] = "new"
            new_tasks.append(current_task)

    for idx, previous_task in enumerate(previous_tasks):
        if idx not in matched_previous_indexes:
            potentially_closed_tasks.append(previous_task)

    return {
        "new_tasks": new_tasks,
        "repeated_tasks": repeated_tasks,
        "potentially_closed_tasks": potentially_closed_tasks,
        "new_count": len(new_tasks),
        "repeated_count": len(repeated_tasks),
        "potentially_closed_count": len(potentially_closed_tasks),
    }

def clean_task_text(text: str) -> str:
    text = strip_speaker_prefix(text)
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\s*(\.\.\.|…)\s*$", "", text)
    return text.strip(" .,!?:;—-")

def expand_task_context(
    segments: list[dict],
    current_index: int,
    current_text: str,
) -> str:
    parts = []

    if current_index - 1 >= 0:
        prev_seg = segments[current_index - 1]
        cur_seg = segments[current_index]

        gap = float(cur_seg.get("start", 0)) - float(prev_seg.get("end", 0))
        prev_text = str(prev_seg.get("text", "")).strip()
        prev_lower = prev_text.lower()

        if gap <= 1.5 and len(prev_text.split()) <= 14:
            if re.search(r"\b(надо|нужно|давайте|тогда|получается|из ближайших задач|здесь надо|лучше)\b", prev_lower):
                parts.append(prev_text)

    parts.append(current_text)

    if current_index + 1 < len(segments):
        next_seg = segments[current_index + 1]
        cur_seg = segments[current_index]

        gap = float(next_seg.get("start", 0)) - float(cur_seg.get("end", 0))
        next_text = str(next_seg.get("text", "")).strip()

        if gap <= 1.5 and len(next_text.split()) <= 14:
            if not next_text.endswith("?"):
                parts.append(next_text)

    result = " ".join(parts)
    result = re.sub(r"\s+", " ", result).strip()
    return clean_task_text(result)


def detect_tasks(segments: list[dict], classified: list[dict], sentiments: list[dict]) -> list[dict]:
    tasks: list[dict] = []

    for idx, (seg, cls, sent) in enumerate(zip(segments, classified, sentiments)):
        text = clean_task_text(seg["text"])
        task_analysis = get_task_analysis_safe(text)

        is_task_by_label = cls.get("label") == "task"
        is_actionable = has_actionable_task_signal(text)

        if not (is_task_by_label and is_actionable):
            continue

        if get_question_kind(text) in {"main", "followup", "echo"}:
            continue

        if text.strip().endswith("?"):
            continue

        if is_casual_imperative(text):
            continue

        if is_filler_heavy(text):
            continue

        task_text = expand_task_context(segments, idx, text)

        # после расширения ещё раз проверяем, что это похоже на задачу
        if not has_actionable_task_signal(task_text):
            task_text = text

        responsible = extract_responsible(task_text)
        deadline = extract_deadline(task_text)

        confidence = max(
            float(cls.get("confidence", 0.0) or 0.0),
            float(task_analysis.get("score", 0.0) or 0.0),
        )

        tasks.append({
            "text": task_text,
            "responsible": responsible,
            "deadline": deadline,
            "confidence": round(confidence, 4),
            "sentiment": sent,
            "status": "new",
            "timecode": f"{seg['start']:.2f} - {seg['end']:.2f} сек.",
            "source": cls.get("source", ""),
            "task_debug": task_analysis,
        })

    return dedupe_tasks(tasks)


def detect_problem_aspects(text: str) -> list[str]:
    patterns = {
        "сроки": [r"\bсрок\b", r"\bдедлайн\b", r"\bзадерж", r"\bдо\s+пятницы\b"],
        "ошибки": [r"\bошиб", r"\bбаг\b", r"\bне\s+работает\b", r"\bсломано\b"],
        "сервер": [r"\bсервер\b", r"\bдеплой\b", r"\bрелиз\b", r"\bинфраструктур"],
        "документы": [r"\bотч[её]т\b", r"\bдоговор\b", r"\bдокумент\b", r"\bпрезентац"],
        "ресурсы": [r"\bресурс", r"\bне\s+успеваем\b", r"\bперегруз"],
    }

    lower = text.lower()
    found = []

    for name, pats in patterns.items():
        if any(re.search(p, lower) for p in pats):
            found.append(name)

    return found


def get_conn():
    return psycopg.connect(DATABASE_URL)


def init_db() -> None:
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS transcriptions (
                id SERIAL PRIMARY KEY,
                filename TEXT NOT NULL,
                original_path TEXT,
                converted_path TEXT,
                duration DOUBLE PRECISION DEFAULT 0,
                language TEXT DEFAULT 'unknown',
                full_text TEXT DEFAULT '',
                segments JSONB DEFAULT '[]',
                tasks_json JSONB DEFAULT '[]',
                qa_json JSONB DEFAULT '[]',
                sentiment_json JSONB DEFAULT '[]',
                analytics_json JSONB DEFAULT '{}',
                meeting_date DATE,
                project_name TEXT DEFAULT '',
                participants TEXT DEFAULT '',
                timing_json JSONB DEFAULT '{}',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                search_vector tsvector
            )
            """
        )

        # Миграции для старой таблицы transcriptions
        cur.execute(
            """
            ALTER TABLE transcriptions
            ADD COLUMN IF NOT EXISTS user_id INTEGER
            """
        )

        cur.execute(
            """
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1
                    FROM pg_constraint
                    WHERE conname = 'transcriptions_user_id_fkey'
                ) THEN
                    ALTER TABLE transcriptions
                    ADD CONSTRAINT transcriptions_user_id_fkey
                    FOREIGN KEY (user_id) REFERENCES users(id);
                END IF;
            END $$;
            """
        )

        cur.execute(
            """
            ALTER TABLE transcriptions
            ADD COLUMN IF NOT EXISTS search_vector tsvector
            """
        )

        cur.execute(
            """
            ALTER TABLE transcriptions
            ADD COLUMN IF NOT EXISTS timing_json JSONB DEFAULT '{}'
            """
        )

        cur.execute(
            """
            ALTER TABLE transcriptions
            ADD COLUMN IF NOT EXISTS participants TEXT DEFAULT ''
            """
        )

        cur.execute(
            """
            ALTER TABLE transcriptions
            ADD COLUMN IF NOT EXISTS project_name TEXT DEFAULT ''
            """
        )

        cur.execute(
            """
            ALTER TABLE transcriptions
            ADD COLUMN IF NOT EXISTS meeting_date DATE
            """
        )

        # Индексы только после того, как колонки точно существуют
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_transcriptions_search
            ON transcriptions USING GIN(search_vector)
            """
        )

        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_transcriptions_user
            ON transcriptions (user_id)
            """
        )

        conn.commit()


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())


def create_token(user_id: int, username: str) -> str:
    payload = {
        "sub": str(user_id),
        "username": username,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=JWT_EXPIRE_MINUTES),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def decode_token(token: str) -> dict:
    return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])


def get_current_user(authorization: str = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Not authenticated")

    token = authorization.split(" ", 1)[1]

    try:
        payload = decode_token(token)
        return {
            "id": int(payload["sub"]),
            "username": payload["username"],
        }
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")
    except Exception:
        raise HTTPException(401, "Invalid token")


def create_user(username: str, password: str) -> int:
    pw_hash = hash_password(password)
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            "INSERT INTO users (username, password_hash) VALUES (%s, %s) RETURNING id",
            (username, pw_hash),
        )
        uid = cur.fetchone()[0]
        conn.commit()
        return uid


def authenticate_user(username: str, password: str) -> Optional[dict]:
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT id, username, password_hash FROM users WHERE username = %s", (username,))
        row = cur.fetchone()
        if not row:
            return None
        if not verify_password(password, row[2]):
            return None
        return {"id": row[0], "username": row[1]}


def update_search_vector(cur, record_id: int) -> None:
    cur.execute(
        """
        UPDATE transcriptions
        SET search_vector =
            to_tsvector(
                'russian',
                coalesce(filename,'') || ' ' ||
                coalesce(project_name,'') || ' ' ||
                coalesce(participants,'') || ' ' ||
                coalesce(full_text,'')
            )
        WHERE id = %s
        """,
        (record_id,),
    )


def get_record(record_id: int, user_id: Optional[int] = None) -> Optional[dict]:
    with get_conn() as conn, conn.cursor() as cur:
        if user_id is not None:
            cur.execute("SELECT * FROM transcriptions WHERE id = %s AND user_id = %s", (record_id, user_id))
        else:
            cur.execute("SELECT * FROM transcriptions WHERE id = %s", (record_id,))
        row = cur.fetchone()

        if not row:
            return None

        cols = [d[0] for d in cur.description]
        return dict(zip(cols, row))


def get_all_records(user_id: Optional[int] = None) -> list[dict]:
    with get_conn() as conn, conn.cursor() as cur:
        if user_id is not None:
            cur.execute("SELECT * FROM transcriptions WHERE user_id = %s ORDER BY created_at DESC", (user_id,))
        else:
            cur.execute("SELECT * FROM transcriptions ORDER BY created_at DESC")
        rows = cur.fetchall()
        cols = [d[0] for d in cur.description]

        return [dict(zip(cols, row)) for row in rows]


def find_previous_record(current_record_id: int, project_name: str = "", user_id: Optional[int] = None) -> Optional[dict]:
    with get_conn() as conn, conn.cursor() as cur:
        if project_name:
            if user_id is not None:
                cur.execute(
                    """
                    SELECT *
                    FROM transcriptions
                    WHERE id < %s AND project_name = %s AND user_id = %s
                    ORDER BY meeting_date DESC NULLS LAST, created_at DESC
                    LIMIT 1
                    """,
                    (current_record_id, project_name, user_id),
                )
            else:
                cur.execute(
                    """
                    SELECT *
                    FROM transcriptions
                    WHERE id < %s AND project_name = %s
                    ORDER BY meeting_date DESC NULLS LAST, created_at DESC
                    LIMIT 1
                    """,
                    (current_record_id, project_name),
                )
        else:
            if user_id is not None:
                cur.execute(
                    """
                    SELECT *
                    FROM transcriptions
                    WHERE id < %s AND user_id = %s
                    ORDER BY meeting_date DESC NULLS LAST, created_at DESC
                    LIMIT 1
                    """,
                    (current_record_id, user_id),
                )
            else:
                cur.execute(
                    """
                    SELECT *
                    FROM transcriptions
                    WHERE id < %s
                    ORDER BY meeting_date DESC NULLS LAST, created_at DESC
                    LIMIT 1
                    """,
                    (current_record_id,),
                )

        row = cur.fetchone()

        if not row:
            return None

        cols = [d[0] for d in cur.description]
        return dict(zip(cols, row))


def search_records(query: str, user_id: Optional[int] = None) -> list[dict]:
    with get_conn() as conn, conn.cursor() as cur:
        if user_id is not None:
            cur.execute(
                """
                SELECT id, filename, project_name, meeting_date, created_at,
                       ts_headline(
                           'russian',
                           full_text,
                           plainto_tsquery('russian', %s),
                           'MaxWords=30, MinWords=10, StartSel=<mark>, StopSel=</mark>'
                       ) AS snippet,
                       ts_rank(search_vector, plainto_tsquery('russian', %s)) AS rank
                FROM transcriptions
                WHERE search_vector @@ plainto_tsquery('russian', %s) AND user_id = %s
                ORDER BY rank DESC, created_at DESC
                LIMIT 30
                """,
                (query, query, query, user_id),
            )
        else:
            cur.execute(
                """
                SELECT id, filename, project_name, meeting_date, created_at,
                       ts_headline(
                           'russian',
                           full_text,
                           plainto_tsquery('russian', %s),
                           'MaxWords=30, MinWords=10, StartSel=<mark>, StopSel=</mark>'
                       ) AS snippet,
                       ts_rank(search_vector, plainto_tsquery('russian', %s)) AS rank
                FROM transcriptions
                WHERE search_vector @@ plainto_tsquery('russian', %s)
                ORDER BY rank DESC, created_at DESC
                LIMIT 30
                """,
                (query, query, query),
            )

        return [
            {
                "id": r[0],
                "filename": r[1],
                "project_name": r[2],
                "meeting_date": r[3],
                "created_at": r[4],
                "snippet": r[5],
                "rank": float(r[6]),
            }
            for r in cur.fetchall()
        ]


def convert_to_wav(input_path: Path, output_path: Path) -> None:
    audio = AudioSegment.from_file(input_path)
    audio = audio.set_channels(1).set_frame_rate(16000)
    audio.export(output_path, format="wav")


def transcribe_audio(wav_path: Path) -> dict:
    kwargs = {
        "beam_size": WHISPER_BEAM_SIZE,
        "vad_filter": True,
        "vad_parameters": {
            "min_silence_duration_ms": 500,
            "speech_pad_ms": 200,
        },
        "condition_on_previous_text": False,
        "temperature": 0,
        "word_timestamps": False,
        "no_speech_threshold": 0.55,
        "compression_ratio_threshold": 2.6,
        "log_prob_threshold": -1.0,
    }

    if WHISPER_LANGUAGE:
        kwargs["language"] = WHISPER_LANGUAGE

    segments_iter, info = whisper_model.transcribe(
        str(wav_path),
        **kwargs,
    )

    segments = []
    text_parts = []
    raw_segment_count = 0

    for seg in segments_iter:
        raw_segment_count += 1
        text = normalize_text(seg.text)

        if not text:
            continue

        split_segments = split_whisper_segment(float(seg.start), float(seg.end), text)
        segments.extend(split_segments)
        text_parts.extend([s["text"] for s in split_segments])

    print(
        f"[transcribe] raw_segments={raw_segment_count}, "
        f"processed_segments={len(segments)}, split={SPLIT_TRANSCRIPT_SEGMENTS}"
    )

    return {
        "language": getattr(info, "language", "unknown") or "unknown",
        "duration": round(getattr(info, "duration", 0.0) or 0.0, 2),
        "segments": segments,
        "text": " ".join(text_parts),
    }


def process_segments(segments: list[dict]) -> tuple[list[dict], list[dict], list[dict], list[dict], dict]:
    texts = [strip_speaker_prefix(s["text"]) for s in segments]
    classified = final_classify(texts)
    sentiments = batch_sentiment(texts)
    tasks = detect_tasks(segments, classified, sentiments)
    qa_pairs = build_qa_pairs_global(segments, classified)

    label_counts: dict[str, int] = {}
    for cls in classified:
        label_counts[cls.get("label", "other")] = label_counts.get(cls.get("label", "other"), 0) + 1

    print(
        f"[nlp] segments={len(segments)}, labels={label_counts}, "
        f"tasks={len(tasks)}, qa_pairs={len(qa_pairs)}, mode={NLP_MODE}"
    )

    sent_values = [s["score"] for s in sentiments]
    avg_sent = round(sum(sent_values) / len(sent_values), 4) if sent_values else 0.0
    neg_ratio = round(
        sum(1 for s in sentiments if s["label"] == "negative") / len(sentiments),
        4,
    ) if sentiments else 0.0

    aspect_counts: dict[str, int] = {}

    for seg in segments:
        for asp in detect_problem_aspects(seg["text"]):
            aspect_counts[asp] = aspect_counts.get(asp, 0) + 1

    analytics = {
        "tasks_count": len(tasks),
        "qa_count": len(qa_pairs),
        "avg_sentiment_score": avg_sent,
        "negative_ratio": neg_ratio,
        "classification_counts": label_counts,
        "top_problem_aspects": sorted(
            [{"aspect": k, "count": v} for k, v in aspect_counts.items()],
            key=lambda x: -x["count"],
        )[:5],
    }

    return classified, sentiments, tasks, qa_pairs, analytics


def build_dynamic_analysis(current_record: dict, previous_record: Optional[dict]) -> dict:
    current_tasks = current_record.get("tasks_json") or []
    current_qa = current_record.get("qa_json") or []
    current_analytics = current_record.get("analytics_json") or {}

    if not previous_record:
        return {
            "has_previous": False,
            "summary": "Предыдущая встреча для сравнения не найдена.",
            "tasks_delta": 0,
            "qa_delta": 0,
            "sentiment_delta": 0,
            "negative_ratio_delta": 0,
            "task_changes": {
                "new_tasks": current_tasks,
                "repeated_tasks": [],
                "potentially_closed_tasks": [],
                "new_count": len(current_tasks),
                "repeated_count": 0,
                "potentially_closed_count": 0,
            },
            "aspect_changes": [],
        }

    previous_tasks = previous_record.get("tasks_json") or []
    previous_qa = previous_record.get("qa_json") or []
    previous_analytics = previous_record.get("analytics_json") or {}

    current_sentiment = float(current_analytics.get("avg_sentiment_score", 0) or 0)
    previous_sentiment = float(previous_analytics.get("avg_sentiment_score", 0) or 0)

    current_negative = float(current_analytics.get("negative_ratio", 0) or 0)
    previous_negative = float(previous_analytics.get("negative_ratio", 0) or 0)

    task_changes = compare_tasks_with_previous(current_tasks, previous_tasks)

    current_aspects = {
        item.get("aspect"): item.get("count", 0)
        for item in current_analytics.get("top_problem_aspects", [])
        if item.get("aspect")
    }

    previous_aspects = {
        item.get("aspect"): item.get("count", 0)
        for item in previous_analytics.get("top_problem_aspects", [])
        if item.get("aspect")
    }

    aspect_changes = []

    for aspect in sorted(set(current_aspects) | set(previous_aspects)):
        current_count = current_aspects.get(aspect, 0)
        previous_count = previous_aspects.get(aspect, 0)

        aspect_changes.append({
            "aspect": aspect,
            "current_count": current_count,
            "previous_count": previous_count,
            "delta": current_count - previous_count,
        })

    tasks_delta = len(current_tasks) - len(previous_tasks)
    qa_delta = len(current_qa) - len(previous_qa)
    sentiment_delta = round(current_sentiment - previous_sentiment, 4)
    negative_ratio_delta = round(current_negative - previous_negative, 4)

    summary_parts = []

    if tasks_delta > 0:
        summary_parts.append(f"Количество задач увеличилось на {tasks_delta}.")
    elif tasks_delta < 0:
        summary_parts.append(f"Количество задач уменьшилось на {abs(tasks_delta)}.")
    else:
        summary_parts.append("Количество задач не изменилось.")

    if qa_delta > 0:
        summary_parts.append(f"Количество пар «вопрос — ответ» увеличилось на {qa_delta}.")
    elif qa_delta < 0:
        summary_parts.append(f"Количество пар «вопрос — ответ» уменьшилось на {abs(qa_delta)}.")
    else:
        summary_parts.append("Количество пар «вопрос — ответ» не изменилось.")

    if sentiment_delta > 0:
        summary_parts.append("Средняя эмоциональная окраска стала более позитивной.")
    elif sentiment_delta < 0:
        summary_parts.append("Средняя эмоциональная окраска стала более негативной.")
    else:
        summary_parts.append("Средняя эмоциональная окраска не изменилась.")

    if negative_ratio_delta > 0:
        summary_parts.append("Доля негативных сегментов выросла.")
    elif negative_ratio_delta < 0:
        summary_parts.append("Доля негативных сегментов снизилась.")
    else:
        summary_parts.append("Доля негативных сегментов осталась прежней.")

    return {
        "has_previous": True,
        "previous_record_id": previous_record.get("id"),
        "previous_filename": previous_record.get("filename"),
        "previous_meeting_date": str(previous_record.get("meeting_date") or ""),
        "tasks_delta": tasks_delta,
        "qa_delta": qa_delta,
        "sentiment_delta": sentiment_delta,
        "negative_ratio_delta": negative_ratio_delta,
        "task_changes": task_changes,
        "aspect_changes": aspect_changes,
        "summary": " ".join(summary_parts),
    }


def save_result_files(stem: str, payload: dict) -> tuple[Path, Path]:
    json_path = RESULTS_DIR / f"{stem}.json"
    txt_path = RESULTS_DIR / f"{stem}.txt"

    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    txt_path.write_text(payload.get("full_text", ""), encoding="utf-8")

    return json_path, txt_path


def store_record(
    filename: str,
    original_path: str,
    converted_path: str,
    transcription: dict,
    tasks: list[dict],
    qa_pairs: list[dict],
    sentiments: list[dict],
    analytics: dict,
    meeting_date: Optional[str],
    project_name: str,
    participants: str,
    timing: dict,
    user_id: Optional[int] = None,
) -> int:
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO transcriptions (
                user_id,
                filename,
                original_path,
                converted_path,
                duration,
                language,
                full_text,
                segments,
                tasks_json,
                qa_json,
                sentiment_json,
                analytics_json,
                meeting_date,
                project_name,
                participants,
                timing_json
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (
                user_id,
                filename,
                original_path,
                converted_path,
                transcription["duration"],
                transcription["language"],
                transcription["text"],
                Json(transcription["segments"]),
                Json(tasks),
                Json(qa_pairs),
                Json(sentiments),
                Json(analytics),
                meeting_date or None,
                project_name,
                participants,
                Json(timing),
            ),
        )

        rid = cur.fetchone()[0]
        update_search_vector(cur, rid)
        conn.commit()

        return rid


def export_excel(record_id: int, user_id: Optional[int] = None) -> Path:
    if not EXCEL_AVAILABLE:
        raise HTTPException(500, "openpyxl not installed")

    rec = get_record(record_id, user_id=user_id)

    if not rec:
        raise HTTPException(404, "Record not found")

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def autosize(ws, max_width: int = 80):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = 12
            for cell in col:
                value = cell.value
                if value is not None:
                    width = max(width, min(len(str(value)) + 2, max_width))
                cell.alignment = wrap
            ws.column_dimensions[letter].width = width

    analytics = rec.get("analytics_json") or {}
    tasks = rec.get("tasks_json") or []
    qa_pairs = rec.get("qa_json") or []
    segments = rec.get("segments") or []
    sentiments = rec.get("sentiment_json") or []
    dynamic = analytics.get("dynamic_analysis") or {}
    timing = rec.get("timing_json") or {}

    # ─── Сводка ─────────────────────────────────────────────

    ws = wb.active
    ws.title = "Сводка"

    summary_rows = [
        ("Файл", rec.get("filename", "")),
        ("Проект", rec.get("project_name") or ""),
        ("Дата встречи", str(rec.get("meeting_date") or "")),
        ("Участники", rec.get("participants") or ""),
        ("Длительность, сек", rec.get("duration") or 0),
        ("Количество задач", len(tasks)),
        ("Количество Q/A", len(qa_pairs)),
        ("Средний sentiment", analytics.get("avg_sentiment_score", 0)),
        ("Негативная доля", analytics.get("negative_ratio", 0)),
        ("Upload, сек", timing.get("upload", "")),
        ("Convert, сек", timing.get("convert", "")),
        ("Transcribe, сек", timing.get("transcribe", "")),
        ("NLP, сек", timing.get("nlp", "")),
        ("Total, сек", timing.get("total", "")),
    ]

    ws.append(["Показатель", "Значение"])
    style_header(ws)

    for row in summary_rows:
        ws.append(list(row))

    for cell in ws["A"]:
        cell.font = bold

    autosize(ws)

    # ─── Динамика ─────────────────────────────────────────────

    dyn_ws = wb.create_sheet("Динамика")
    dyn_ws.append(["Показатель", "Значение"])
    style_header(dyn_ws)

    dyn_rows = [
        ("Есть предыдущая встреча", "Да" if dynamic.get("has_previous") else "Нет"),
        ("Предыдущий файл", dynamic.get("previous_filename", "")),
        ("Дата предыдущей встречи", dynamic.get("previous_meeting_date", "")),
        ("Изменение количества задач", dynamic.get("tasks_delta", 0)),
        ("Изменение количества Q/A", dynamic.get("qa_delta", 0)),
        ("Изменение sentiment", dynamic.get("sentiment_delta", 0)),
        ("Изменение негативной доли", dynamic.get("negative_ratio_delta", 0)),
        ("Вывод", dynamic.get("summary", "")),
    ]

    for row in dyn_rows:
        dyn_ws.append(list(row))

    autosize(dyn_ws)

    # ─── Задачи ─────────────────────────────────────────────

    tasks_ws = wb.create_sheet("Задачи")
    tasks_ws.append([
        "№",
        "Задача",
        "Ответственный",
        "Срок",
        "Confidence",
        "Sentiment",
        "Timecode",
    ])
    style_header(tasks_ws)

    if tasks:
        for i, task in enumerate(tasks, 1):
            sent = task.get("sentiment") or {}
            tasks_ws.append([
                i,
                task.get("text", ""),
                task.get("responsible") or "",
                task.get("deadline") or "",
                task.get("confidence", ""),
                sent.get("label", ""),
                task.get("timecode", ""),
            ])
    else:
        tasks_ws.append(["", "Задачи не найдены", "", "", "", "", ""])

    autosize(tasks_ws, max_width=100)

    # ─── Q/A ─────────────────────────────────────────────

    qa_ws = wb.create_sheet("Вопросы-ответы")
    qa_ws.append([
        "№",
        "Вопрос",
        "Ответ",
        "Время вопроса",
        "Время ответа",
        "Confidence",
    ])
    style_header(qa_ws)

    if qa_pairs:
        for i, qa in enumerate(qa_pairs, 1):
            qa_ws.append([
                i,
                qa.get("question", ""),
                qa.get("answer", ""),
                qa.get("question_timecode", ""),
                qa.get("answer_timecode", ""),
                qa.get("confidence", ""),
            ])
    else:
        qa_ws.append(["", "Q/A не найдены", "", "", "", ""])

    qa_ws.column_dimensions["B"].width = 70
    qa_ws.column_dimensions["C"].width = 100

    for row in qa_ws.iter_rows():
        for cell in row:
            cell.alignment = wrap

    # ─── Транскрибация полным текстом ─────────────────────

    text_ws = wb.create_sheet("Транскрибация")
    text_ws.append(["Полная транскрибация"])
    style_header(text_ws)

    full_text = rec.get("full_text") or ""
    text_ws.append([full_text if full_text else "Транскрибация не найдена"])
    text_ws.column_dimensions["A"].width = 140

    for row in text_ws.iter_rows():
        for cell in row:
            cell.alignment = wrap

    # ─── Сегменты ─────────────────────────────────────────

    seg_ws = wb.create_sheet("Сегменты")
    seg_ws.append([
        "№",
        "Время",
        "Спикер",
        "Текст",
        "Тип",
        "Confidence",
        "Sentiment",
        "Sentiment score",
    ])
    style_header(seg_ws)

    if segments:
        for i, seg in enumerate(segments, 1):
            seg_ws.append([
                i,
                seg.get("timecode") or f"[{seg.get('start', '')} - {seg.get('end', '')} сек.]",
                seg.get("speaker", ""),
                seg.get("text", ""),
                seg.get("predicted_label", ""),
                seg.get("prediction_confidence", ""),
                seg.get("sentiment_label", ""),
                seg.get("sentiment_score", ""),
            ])
    else:
        seg_ws.append(["", "", "", "Сегменты не найдены", "", "", "", ""])

    seg_ws.column_dimensions["D"].width = 120

    for row in seg_ws.iter_rows():
        for cell in row:
            cell.alignment = wrap

    # ─── Полезная аналитика ───────────────────────────────

    useful_ws = wb.create_sheet("Полезное")
    useful_ws.append(["Раздел", "Данные"])
    style_header(useful_ws)

    useful_ws.append(["Краткий вывод", dynamic.get("summary") or "Нет данных для динамического анализа."])

    aspects = analytics.get("top_problem_aspects") or []
    if aspects:
        for item in aspects:
            useful_ws.append([
                "Проблемный аспект",
                f"{item.get('aspect', '')}: {item.get('count', 0)}",
            ])
    else:
        useful_ws.append(["Проблемные аспекты", "Не найдены"])

    useful_ws.append(["Рекомендация", "Проверить Q/A пары вручную, если встреча длинная или в ней несколько тем подряд."])
    useful_ws.append(["Рекомендация", "Для задач лучше использовать рабочие встречи, где есть фразы: нужно, сделай, ответственный, срок, дедлайн."])

    autosize(useful_ws, max_width=100)

    out = EXPORTS_DIR / f"report_{record_id}.xlsx"
    wb.save(out)

    return out


def export_pdf(record_id: int, user_id: Optional[int] = None) -> Path:
    if not PDF_AVAILABLE:
        raise HTTPException(500, "fpdf/fpdf2 not installed")

    rec = get_record(record_id, user_id=user_id)

    if not rec:
        raise HTTPException(404, "Record not found")

    analytics = rec.get("analytics_json") or {}
    tasks = rec.get("tasks_json") or []
    qa_pairs = rec.get("qa_json") or []
    segments = rec.get("segments") or []
    dynamic = analytics.get("dynamic_analysis") or {}
    timing = rec.get("timing_json") or {}

    pdf = FPDF()
    pdf.set_margins(12, 12, 12)
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    font_path_candidates = [
        "DejaVuSans.ttf",
        "fonts/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]

    font_path = None
    for candidate in font_path_candidates:
        if Path(candidate).exists():
            font_path = candidate
            break

    if font_path:
        pdf.add_font("MainFont", "", font_path)
        pdf.set_font("MainFont", size=10)
    else:
        # Важно: без unicode-шрифта русский PDF может ломаться.
        raise HTTPException(
            500,
            "Unicode font not found. Put DejaVuSans.ttf into project root or fonts/DejaVuSans.ttf"
        )

    def safe_text(value) -> str:
        text = "" if value is None else str(value)
        text = text.replace("\r", "")
        text = text.replace("\t", " ")
        text = text.replace("—", "-")
        text = text.replace("–", "-")
        text = text.replace("•", "-")
        return text

    def pdf_width() -> float:
        return pdf.w - pdf.l_margin - pdf.r_margin

    def write_multicell(text: str, h: int = 5, size: int = 10):
        text = safe_text(text)

        if not text.strip():
            text = "-"

        pdf.set_font("MainFont", size=size)

        wrapped = textwrap.fill(
            text,
            width=90,
            break_long_words=True,
            break_on_hyphens=True,
            replace_whitespace=False,
        )

        try:
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(
                w=pdf_width(),
                h=h,
                text=wrapped,
                new_x="LMARGIN",
                new_y="NEXT",
            )
        except TypeError:
            # fallback для старых версий fpdf
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(pdf_width(), h, wrapped)
        except FPDFException:
            # аварийный fallback, если попалось что-то совсем неудобное для переноса
            for i in range(0, len(text), 60):
                pdf.set_x(pdf.l_margin)
                pdf.multi_cell(pdf_width(), h, text[i:i + 60])

    def title(text: str):
        write_multicell(text, h=8, size=14)
        pdf.ln(2)


    def subtitle(text: str):
        pdf.ln(3)
        write_multicell(text, h=7, size=12)
    def line(label, value):
        write_multicell(f"{safe_text(label)}: {safe_text(value)}", h=6, size=10)
    def paragraph(text: str, h: int = 5):
        write_multicell(text, h=h, size=10)
        pdf.ln(1)

    title(f"PM Insights - отчет по встрече")

    line("Файл", rec.get("filename", ""))
    line("Проект", rec.get("project_name") or "")
    line("Дата", rec.get("meeting_date") or "")
    line("Участники", rec.get("participants") or "")
    line("Длительность, сек", rec.get("duration") or 0)
    line("Количество задач", len(tasks))
    line("Количество Q/A", len(qa_pairs))
    line("Средний sentiment", analytics.get("avg_sentiment_score", 0))
    line("Негативная доля", analytics.get("negative_ratio", 0))
    line("Общее время обработки, сек", timing.get("total", ""))

    subtitle("Динамический анализ")
    paragraph(dynamic.get("summary") or "Нет данных для динамического анализа.")

    subtitle("Задачи")
    if tasks:
        for i, task in enumerate(tasks, 1):
            paragraph(
                f"{i}. {task.get('text', '')}\n"
                f"Ответственный: {task.get('responsible') or '-'} | "
                f"Срок: {task.get('deadline') or '-'} | "
                f"Confidence: {task.get('confidence', '')} | "
                f"Timecode: {task.get('timecode', '')}"
            )
    else:
        paragraph("Задачи не найдены.")

    subtitle("Вопросы - ответы")
    if qa_pairs:
        for i, qa in enumerate(qa_pairs, 1):
            paragraph(
                f"{i}. Вопрос: {qa.get('question', '')}\n"
                f"Время вопроса: {qa.get('question_timecode', '')}\n"
                f"Ответ: {qa.get('answer', '')}\n"
                f"Время ответа: {qa.get('answer_timecode', '')}\n"
                f"Confidence: {qa.get('confidence', '')}"
            )
    else:
        paragraph("Q/A пары не найдены.")

    subtitle("Полная транскрибация")
    full_text = rec.get("full_text") or ""

    if full_text:
        # Чтобы PDF не падал на огромном тексте, режем на куски.
        chunks = [full_text[i:i + 2500] for i in range(0, len(full_text), 2500)]
        for chunk in chunks:
            paragraph(chunk, h=5)
    else:
        paragraph("Транскрибация не найдена.")

    subtitle("Сегменты транскрибации")
    if segments:
        for i, seg in enumerate(segments, 1):
            paragraph(
                f"{i}. {seg.get('timecode') or f'[{seg.get('start', '')} - {seg.get('end', '')} сек.]'} "
                f"{seg.get('speaker', '')}\n"
                f"{seg.get('text', '')}\n"
                f"Тип: {seg.get('predicted_label', '')} | "
                f"Sentiment: {seg.get('sentiment_label', '')} "
                f"({seg.get('sentiment_score', '')})"
            )
    else:
        paragraph("Сегменты не найдены.")

    out = EXPORTS_DIR / f"report_{record_id}.pdf"
    pdf.output(str(out))

    return out


def page(title: str, content: str, active: str = "home") -> HTMLResponse:
    nav = f"""
    <nav class='nav'>
      <a class='{ 'active' if active == 'home' else '' }' href='/'>Загрузка</a>
      <a class='{ 'active' if active == 'analytics' else '' }' href='/analytics/'>Аналитика</a>
      <a class='{ 'active' if active == 'search' else '' }' href='/search/'>Поиск</a>
    </nav>
    """

    html_doc = f"""
    <!doctype html>
    <html lang='ru'>
    <head>
      <meta charset='utf-8'>
      <meta name='viewport' content='width=device-width, initial-scale=1'>
      <title>{html.escape(title)}</title>
      <script src='https://cdn.jsdelivr.net/npm/chart.js'></script>
      <style>
        body {{ font-family: Arial, sans-serif; margin: 0; background: #f5f7fb; color: #1f2937; }}
        .wrap {{ max-width: 1180px; margin: 0 auto; padding: 20px; }}
        .nav {{ display:flex; gap:16px; padding:14px 20px; background:#111827; }}
        .nav a {{ color:#cbd5e1; text-decoration:none; font-weight:600; }}
        .nav a.active {{ color:#fff; }}
        .cards {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(190px,1fr)); gap:14px; margin:16px 0; }}
        .card, .section {{ background:#fff; border-radius:12px; padding:16px; box-shadow:0 2px 8px rgba(0,0,0,.05); }}
        .card h4 {{ margin:0 0 8px; color:#6b7280; font-size:13px; }}
        .val {{ font-size:28px; font-weight:700; }}
        .sub {{ font-size:12px; color:#6b7280; margin-top:6px; }}
        .btn {{ display:inline-block; padding:10px 14px; border-radius:8px; text-decoration:none; color:#fff; border:none; cursor:pointer; }}
        .btn-blue {{ background:#2563eb; }}
        .btn-green {{ background:#16a34a; }}
        .btn-red {{ background:#dc2626; }}
        .btn-gray {{ background:#6b7280; }}
        .row {{ display:flex; gap:10px; flex-wrap:wrap; }}
        input[type=text], input[type=date], textarea {{ width:100%; padding:10px 12px; border:1px solid #d1d5db; border-radius:8px; box-sizing:border-box; }}
        label {{ font-weight:600; font-size:14px; display:block; margin-bottom:6px; }}
        .task-card {{ border:1px solid #e5e7eb; border-radius:10px; padding:12px; margin:10px 0; background:#fff; }}
        .seg {{ padding:10px 12px; border-left:6px solid #d1d5db; background:#f9fafb; margin:8px 0; border-radius:8px; }}
        .seg-pos {{ border-color:#10b981; }}
        .seg-neg {{ border-color:#ef4444; }}
        .seg-neu {{ border-color:#9ca3af; }}
        .timing {{ font-size:14px; color:#4b5563; display:flex; gap:16px; flex-wrap:wrap; margin:12px 0; }}
        table {{ width:100%; border-collapse:collapse; }}
        th, td {{ padding:10px; border-bottom:1px solid #e5e7eb; text-align:left; vertical-align:top; }}
        mark {{ background:#fde68a; padding:0 2px; }}
        .muted {{ color:#6b7280; }}
      </style>
    </head>
    <body>
      {nav}
      <div class='wrap'>
        {content}
      </div>
    </body>
    </html>
    """

    return HTMLResponse(html_doc)


@app.get("/favicon.ico")
async def favicon() -> Response:
    return Response(status_code=204)


@app.on_event("startup")
async def startup_event() -> None:
    init_db()


# ─── JSON API: Auth ───────────────────────────────────────────────────────────

from pydantic import BaseModel


class RegisterRequest(BaseModel):
    username: str
    password: str


class LoginRequest(BaseModel):
    username: str
    password: str


@app.post("/api/auth/register")
async def api_register(body: RegisterRequest):
    if len(body.username) < 3:
        raise HTTPException(400, "Username must be at least 3 characters")
    if len(body.password) < 4:
        raise HTTPException(400, "Password must be at least 4 characters")
    try:
        uid = create_user(body.username, body.password)
    except psycopg.errors.UniqueViolation:
        raise HTTPException(409, "Username already taken")
    token = create_token(uid, body.username)
    return {"token": token, "user": {"id": uid, "username": body.username}}


@app.post("/api/auth/login")
async def api_login(body: LoginRequest):
    user = authenticate_user(body.username, body.password)
    if not user:
        raise HTTPException(401, "Invalid username or password")
    token = create_token(user["id"], user["username"])
    return {"token": token, "user": user}


@app.get("/api/auth/me")
async def api_me(user: dict = Depends(get_current_user)):
    return {"user": user}


# ─── JSON API: Records (authenticated) ─────────────────────────────────────

def _serialize_record(rec: dict) -> dict:
    from datetime import date as _date
    out = {}
    for k, v in rec.items():
        if k == "search_vector":
            continue
        if isinstance(v, _date):
            out[k] = v.isoformat()
        elif isinstance(v, datetime):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out


@app.get("/api/records")
async def api_records(user: dict = Depends(get_current_user)):
    records = get_all_records(user_id=user["id"])
    return [_serialize_record(r) for r in records]


@app.get("/api/records/{record_id}")
async def api_record_detail(record_id: int, user: dict = Depends(get_current_user)):
    rec = get_record(record_id, user_id=user["id"])
    if not rec:
        raise HTTPException(404, "Record not found")
    return _serialize_record(rec)


@app.post("/api/upload")
async def api_upload_file(
    file: UploadFile = File(...),
    meeting_date: Optional[str] = Form(None),
    project_name: str = Form(""),
    participants: str = Form(""),
    user: dict = Depends(get_current_user),
):
    ext = Path(file.filename or "").suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Only these formats are supported: {', '.join(sorted(ALLOWED_EXTENSIONS))}")

    safe_name = Path(file.filename).name
    file_hash = hashlib.md5(f"{safe_name}_{time.time()}".encode()).hexdigest()[:12]

    original_path = UPLOAD_DIR / f"{file_hash}_{safe_name}"
    wav_path = CONVERTED_DIR / f"{file_hash}.wav"

    t0 = time.perf_counter()
    with original_path.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    t1 = time.perf_counter()

    convert_to_wav(original_path, wav_path)
    t2 = time.perf_counter()

    transcription = transcribe_audio(wav_path)
    transcription["segments"] = apply_speaker_diarization(wav_path, transcription["segments"])
    transcription["text"] = " ".join(
    s.get("text", "") for s in transcription["segments"] if s.get("text")
    )
    t3 = time.perf_counter()

    classified, sentiments, tasks, qa_pairs, analytics = process_segments(transcription["segments"])

    t4 = time.perf_counter()

    enriched_segments = []
    for seg, cls, sent in zip(transcription["segments"], classified, sentiments):
        enriched_segments.append({
            **seg,
            "predicted_label": cls.get("label", "other"),
            "prediction_confidence": cls.get("confidence", 0),
            "prediction_source": cls.get("source", "rules"),
            "prediction_debug": cls.get("debug", {}),
            "task_debug": cls.get("task_debug", {}),
            "sentiment_label": sent.get("label", "neutral"),
            "sentiment_score": sent.get("score", 0),
            "timecode": f"[{seg['start']:.2f} - {seg['end']:.2f} сек.]",
        })

    timing = {
        "upload": round(t1 - t0, 2),
        "convert": round(t2 - t1, 2),
        "transcribe": round(t3 - t2, 2),
        "nlp": round(t4 - t3, 2),
        "total": round(t4 - t0, 2),
    }

    payload = {
        "filename": safe_name,
        "project_name": project_name,
        "participants": participants,
        "meeting_date": meeting_date,
        "duration": transcription["duration"],
        "language": transcription["language"],
        "full_text": transcription["text"],
        "segments": enriched_segments,
        "tasks": tasks,
        "qa_pairs": qa_pairs,
        "analytics": analytics,
        "timing": timing,
    }

    stem = original_path.stem
    save_result_files(stem, payload)

    record_id = store_record(
        filename=safe_name,
        original_path=str(original_path),
        converted_path=str(wav_path),
        transcription={**transcription, "segments": enriched_segments},
        tasks=tasks,
        qa_pairs=qa_pairs,
        sentiments=sentiments,
        analytics=analytics,
        meeting_date=meeting_date,
        project_name=project_name,
        participants=participants,
        timing=timing,
        user_id=user["id"],
    )

    current_record = get_record(record_id)
    previous_record = find_previous_record(record_id, project_name, user_id=user["id"])

    if current_record:
        dynamic_analysis = build_dynamic_analysis(current_record, previous_record)
        analytics["dynamic_analysis"] = dynamic_analysis
        payload["analytics"] = analytics
        save_result_files(stem, payload)

        with get_conn() as conn, conn.cursor() as cur:
            cur.execute(
                "UPDATE transcriptions SET analytics_json = %s WHERE id = %s",
                (Json(analytics), record_id),
            )
            conn.commit()

    return {"ok": True, "record_id": record_id}


@app.delete("/api/records/{record_id}")
async def api_delete_record(record_id: int, user: dict = Depends(get_current_user)):
    rec = get_record(record_id, user_id=user["id"])
    if not rec:
        raise HTTPException(404, "Record not found")
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM transcriptions WHERE id = %s AND user_id = %s", (record_id, user["id"]))
        conn.commit()
    return {"ok": True}


@app.get("/api/search")
async def api_search(q: str = "", user: dict = Depends(get_current_user)):
    if not q:
        return []
    results = search_records(q, user_id=user["id"])
    for r in results:
        if r.get("meeting_date"):
            r["meeting_date"] = r["meeting_date"].isoformat() if hasattr(r["meeting_date"], "isoformat") else str(r["meeting_date"])
        if r.get("created_at"):
            r["created_at"] = r["created_at"].isoformat() if hasattr(r["created_at"], "isoformat") else str(r["created_at"])
    return results


@app.get("/api/export/excel/{record_id}")
async def api_excel_export(record_id: int, user: dict = Depends(get_current_user)):
    rec = get_record(record_id, user_id=user["id"])
    if not rec:
        raise HTTPException(404, "Record not found")

    path = export_excel(record_id, user_id=user["id"])
    safe_filename = re.sub(r'[^\w\-.]', '_', rec.get("filename", f"record_{record_id}"))

    return FileResponse(
        path,
        filename=f"report_{safe_filename}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/export/pdf/{record_id}")
async def api_pdf_export(record_id: int, user: dict = Depends(get_current_user)):
    rec = get_record(record_id, user_id=user["id"])
    if not rec:
        raise HTTPException(404, "Record not found")

    path = export_pdf(record_id, user_id=user["id"])
    safe_filename = re.sub(r'[^\w\-.]', '_', rec.get("filename", f"record_{record_id}"))

    return FileResponse(
        path,
        filename=f"report_{safe_filename}.pdf",
        media_type="application/pdf",
    )


# ─── Legacy HTML pages (kept for backward compatibility) ──────────────────────

@app.get("/", response_class=HTMLResponse)
async def home() -> HTMLResponse:
    records = get_all_records()[:15]
    rows = ""

    for rec in records:
        analytics = rec.get("analytics_json") or {}
        timing = rec.get("timing_json") or {}

        rows += f"""
        <tr>
            <td><a href='/result/{rec['id']}'>{html.escape(rec['filename'])}</a></td>
            <td>{html.escape(rec.get('project_name') or '')}</td>
            <td>{rec.get('meeting_date') or ''}</td>
            <td>{round(rec.get('duration') or 0, 2)}</td>
            <td>{len(rec.get('tasks_json') or [])}</td>
            <td>{len(rec.get('qa_json') or [])}</td>
            <td>{analytics.get('avg_sentiment_score', 0)}</td>
            <td>{round(float((timing or {}).get('total', 0)), 2)}</td>
        </tr>
        """

    content = f"""
    <div class='section'>
      <h2>PM Insights</h2>
      <p class='muted'>Текущий режим NLP: <b>{html.escape(NLP_MODE)}</b></p>
      <form action='/upload/' method='post' enctype='multipart/form-data'>
        <div class='row'>
          <div style='flex:2'>
            <label>Аудиофайл</label>
            <input type='file' name='file' required>
          </div>
          <div style='flex:1'>
            <label>Дата встречи</label>
            <input type='date' name='meeting_date'>
          </div>
        </div>
        <div class='row' style='margin-top:10px'>
          <div style='flex:1'>
            <label>Проект</label>
            <input type='text' name='project_name' placeholder='Например, PM Insights'>
          </div>
          <div style='flex:1'>
            <label>Участники</label>
            <input type='text' name='participants' placeholder='Иван, Анна, Сергей'>
          </div>
        </div>
        <div style='margin-top:14px'>
          <button class='btn btn-blue' type='submit'>Загрузить и обработать</button>
        </div>
      </form>
    </div>
    <div class='section'>
      <h3>История обработанных встреч</h3>
      <div style='overflow-x:auto'>
        <table>
          <thead>
            <tr>
              <th>Файл</th>
              <th>Проект</th>
              <th>Дата</th>
              <th>Длительность</th>
              <th>Задачи</th>
              <th>Q/A</th>
              <th>Sentiment</th>
              <th>Общее время</th>
            </tr>
          </thead>
          <tbody>
            {rows or "<tr><td colspan='8' class='muted'>Пока нет данных</td></tr>"}
          </tbody>
        </table>
      </div>
    </div>
    """

    return page("PM Insights", content, "home")


@app.post("/upload/")
async def upload_file(
    file: UploadFile = File(...),
    meeting_date: Optional[str] = Form(None),
    project_name: str = Form(""),
    participants: str = Form(""),
):
    ext = Path(file.filename or "").suffix.lower()

    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Поддерживаются только: {', '.join(sorted(ALLOWED_EXTENSIONS))}")

    safe_name = Path(file.filename).name
    file_hash = hashlib.md5(f"{safe_name}_{time.time()}".encode()).hexdigest()[:12]

    original_path = UPLOAD_DIR / f"{file_hash}_{safe_name}"
    wav_path = CONVERTED_DIR / f"{file_hash}.wav"

    t0 = time.perf_counter()

    with original_path.open("wb") as f:
        shutil.copyfileobj(file.file, f)

    t1 = time.perf_counter()

    convert_to_wav(original_path, wav_path)

    t2 = time.perf_counter()

    transcription = transcribe_audio(wav_path)
    transcription["segments"] = apply_speaker_diarization(wav_path, transcription["segments"])
    transcription["text"] = " ".join(
        s.get("text", "") for s in transcription["segments"] if s.get("text")
    )
    t3 = time.perf_counter()

    classified, sentiments, tasks, qa_pairs, analytics = process_segments(transcription["segments"])

    t4 = time.perf_counter()

    enriched_segments = []

    for seg, cls, sent in zip(transcription["segments"], classified, sentiments):
        enriched_segments.append({
            **seg,
            "predicted_label": cls.get("label", "other"),
            "prediction_confidence": cls.get("confidence", 0),
            "prediction_source": cls.get("source", "rules"),
            "prediction_debug": cls.get("debug", {}),
            "task_debug": cls.get("task_debug", {}),
            "sentiment_label": sent.get("label", "neutral"),
            "sentiment_score": sent.get("score", 0),
            "timecode": f"[{seg['start']:.2f} - {seg['end']:.2f} сек.]",
        })

    timing = {
        "upload": round(t1 - t0, 2),
        "convert": round(t2 - t1, 2),
        "transcribe": round(t3 - t2, 2),
        "nlp": round(t4 - t3, 2),
        "total": round(t4 - t0, 2),
    }

    payload = {
        "filename": safe_name,
        "project_name": project_name,
        "participants": participants,
        "meeting_date": meeting_date,
        "duration": transcription["duration"],
        "language": transcription["language"],
        "full_text": transcription["text"],
        "segments": enriched_segments,
        "tasks": tasks,
        "qa_pairs": qa_pairs,
        "analytics": analytics,
        "timing": timing,
    }

    stem = original_path.stem
    save_result_files(stem, payload)

    record_id = store_record(
        filename=safe_name,
        original_path=str(original_path),
        converted_path=str(wav_path),
        transcription={**transcription, "segments": enriched_segments},
        tasks=tasks,
        qa_pairs=qa_pairs,
        sentiments=sentiments,
        analytics=analytics,
        meeting_date=meeting_date,
        project_name=project_name,
        participants=participants,
        timing=timing,
    )

    current_record = get_record(record_id)
    previous_record = find_previous_record(record_id, project_name)

    if current_record:
        dynamic_analysis = build_dynamic_analysis(current_record, previous_record)
        analytics["dynamic_analysis"] = dynamic_analysis
        payload["analytics"] = analytics
        save_result_files(stem, payload)

        with get_conn() as conn, conn.cursor() as cur:
            cur.execute(
                """
                UPDATE transcriptions
                SET analytics_json = %s
                WHERE id = %s
                """,
                (Json(analytics), record_id),
            )
            conn.commit()

    return JSONResponse({
        "ok": True,
        "record_id": record_id,
        "redirect": f"/result/{record_id}",
    })


@app.get("/result/{record_id}", response_class=HTMLResponse)
async def result_page(record_id: int) -> HTMLResponse:
    rec = get_record(record_id)

    if not rec:
        raise HTTPException(404, "Запись не найдена")

    tasks = rec.get("tasks_json") or []
    qa_pairs = rec.get("qa_json") or []
    segments = rec.get("segments") or []
    analytics = rec.get("analytics_json") or {}
    timing = rec.get("timing_json") or {}

    dynamic = analytics.get("dynamic_analysis") or {}
    task_changes = dynamic.get("task_changes") or {}

    new_tasks = task_changes.get("new_tasks") or []
    repeated_tasks = task_changes.get("repeated_tasks") or []
    closed_tasks = task_changes.get("potentially_closed_tasks") or []
    aspect_changes = dynamic.get("aspect_changes") or []

    new_tasks_html = "".join(
        f"<div class='task-card'><b>{html.escape(t.get('text', ''))}</b>"
        f"<div class='sub'>Ответственный: {html.escape(t.get('responsible') or '—')} | "
        f"Срок: {html.escape(t.get('deadline') or '—')}</div></div>"
        for t in new_tasks
    )

    repeated_tasks_html = "".join(
        f"<div class='task-card'><b>Текущая:</b> {html.escape((item.get('current_task') or {}).get('text', ''))}<br>"
        f"<b>Похожая из прошлой встречи:</b> {html.escape((item.get('previous_task') or {}).get('text', ''))}"
        f"<div class='sub'>Сходство: {item.get('similarity', 0)}</div></div>"
        for item in repeated_tasks
    )

    closed_tasks_html = "".join(
        f"<div class='task-card'><b>{html.escape(t.get('text', ''))}</b>"
        f"<div class='sub'>Задача была в прошлой встрече, но не найдена в текущей.</div></div>"
        for t in closed_tasks
    )

    aspect_rows = "".join(
        f"<tr>"
        f"<td>{html.escape(a.get('aspect', ''))}</td>"
        f"<td>{a.get('previous_count', 0)}</td>"
        f"<td>{a.get('current_count', 0)}</td>"
        f"<td>{a.get('delta', 0)}</td>"
        f"</tr>"
        for a in aspect_changes
    )

    dynamic_html = f"""
    <div class='section'>
      <h3>Динамический анализ изменений</h3>
      <p>{html.escape(dynamic.get('summary', 'Нет данных для динамического анализа.'))}</p>

      <div class='cards'>
        <div class='card'><h4>Изменение задач</h4><div class='val'>{dynamic.get('tasks_delta', 0)}</div></div>
        <div class='card'><h4>Изменение Q/A</h4><div class='val'>{dynamic.get('qa_delta', 0)}</div></div>
        <div class='card'><h4>Изменение sentiment</h4><div class='val' style='font-size:22px'>{dynamic.get('sentiment_delta', 0)}</div></div>
        <div class='card'><h4>Изменение негатива</h4><div class='val' style='font-size:22px'>{dynamic.get('negative_ratio_delta', 0)}</div></div>
      </div>

      <h4>Новые задачи ({len(new_tasks)})</h4>
      {new_tasks_html or '<p class="muted">Новые задачи не найдены</p>'}

      <h4>Повторяющиеся задачи ({len(repeated_tasks)})</h4>
      {repeated_tasks_html or '<p class="muted">Повторяющиеся задачи не найдены</p>'}

      <h4>Потенциально закрытые задачи ({len(closed_tasks)})</h4>
      {closed_tasks_html or '<p class="muted">Потенциально закрытые задачи не найдены</p>'}

      <h4>Изменение проблемных аспектов</h4>
      <table>
        <thead>
          <tr>
            <th>Аспект</th>
            <th>Прошлая встреча</th>
            <th>Текущая встреча</th>
            <th>Изменение</th>
          </tr>
        </thead>
        <tbody>
          {aspect_rows or "<tr><td colspan='4' class='muted'>Нет данных</td></tr>"}
        </tbody>
      </table>
    </div>
    """

    tasks_html = "".join(
        f"<div class='task-card'><b>{i}. {html.escape(t.get('text', ''))}</b><div class='sub'>"
        f"Ответственный: <b>{html.escape(t.get('responsible') or '—')}</b> | "
        f"Срок: <b>{html.escape(t.get('deadline') or '—')}</b> | "
        f"Confidence: {t.get('confidence', 0)} | "
        f"Sentiment: {(t.get('sentiment') or {}).get('label', 'neutral')} | "
        f"{t.get('timecode', '')}</div>"
        f"<div class='sub'>Rule reasons: {html.escape(str((t.get('task_debug') or {}).get('reasons', [])))}</div></div>"
        for i, t in enumerate(tasks, 1)
    )

    qa_html = "".join(
        f"<div class='task-card'><b>В:</b> {html.escape(q.get('question', ''))} "
        f"<span class='muted'>{q.get('question_timecode', '')}</span><br>"
        f"<b>О:</b> {html.escape(q.get('answer', ''))} "
        f"<span class='muted'>{q.get('answer_timecode', '')}</span>"
        f"<div class='sub'>Confidence: {q.get('confidence', 0)}</div></div>"
        for q in qa_pairs
    )

    segments_html = ""
    chart_labels = []
    chart_scores = []

    for seg in segments:
        sent_label = seg.get("sentiment_label", "neutral")
        seg_class = "seg-neu"

        if sent_label == "positive":
            seg_class = "seg-pos"
        elif sent_label == "negative":
            seg_class = "seg-neg"

        chart_labels.append(seg.get("timecode", "")[:12])
        chart_scores.append(seg.get("sentiment_score", 0))

        segments_html += (
            f"<div class='seg {seg_class}'>"
            f"<span class='muted'>{seg.get('timecode', '')}</span> "
            f"<b>{html.escape(seg.get('speaker', 'SPEAKER_00'))}:</b> "
            f"{html.escape(seg.get('text', ''))}"
            f"<span style='float:right' class='muted'>{html.escape(sent_label)} ({seg.get('sentiment_score', 0)})</span>"
            f"<br><span class='muted'>pred: {html.escape(seg.get('predicted_label', 'other'))}, "
            f"conf={seg.get('prediction_confidence', 0)}, "
            f"src={html.escape(seg.get('prediction_source', 'rules'))}</span></div>"
        )

    content = f"""
    <div class='cards'>
      <div class='card'>
        <h4>Файл</h4>
        <div style='font-weight:700'>{html.escape(rec['filename'])}</div>
        <div class='sub'>Проект: {html.escape(rec.get('project_name') or '—')}</div>
      </div>
      <div class='card'>
        <h4>Дата</h4>
        <div class='val' style='font-size:22px'>{rec.get('meeting_date') or '—'}</div>
        <div class='sub'>Участники: {html.escape(rec.get('participants') or '—')}</div>
      </div>
      <div class='card'><h4>Длительность</h4><div class='val'>{rec.get('duration', 0)}<span style='font-size:14px'>с</span></div></div>
      <div class='card'><h4>Задач</h4><div class='val'>{len(tasks)}</div></div>
      <div class='card'><h4>Q/A пар</h4><div class='val'>{len(qa_pairs)}</div></div>
      <div class='card'>
        <h4>Sentiment</h4>
        <div class='val' style='font-size:20px'>{analytics.get('avg_sentiment_score', 0)}</div>
        <div class='sub'>Негатив: {analytics.get('negative_ratio', 0)}</div>
      </div>
    </div>

    <div class='timing'>
      <span>upload: {timing.get('upload', 0)}с</span>
      <span>convert: {timing.get('convert', 0)}с</span>
      <span>transcribe: {timing.get('transcribe', 0)}с</span>
      <span>nlp: {timing.get('nlp', 0)}с</span>
      <span>total: {timing.get('total', 0)}с</span>
    </div>

    <div class='row' style='margin:10px 0 18px'>
      <a class='btn btn-red' href='/export/pdf/{record_id}'>PDF</a>
      <a class='btn btn-green' href='/export/excel/{record_id}'>Excel</a>
      <a class='btn btn-blue' href='/results/{Path(rec['original_path']).stem}.json' target='_blank'>JSON</a>
      <a class='btn btn-gray' href='/'>← Назад</a>
    </div>

    {dynamic_html}

    <div class='section'><h3>Задачи ({len(tasks)})</h3>{tasks_html or '<p class="muted">Задачи не найдены</p>'}</div>
    <div class='section'><h3>Вопросы — ответы ({len(qa_pairs)})</h3>{qa_html or '<p class="muted">Q/A не найдены</p>'}</div>
    <div class='section'><h3>Транскрипция с тональностью</h3>{segments_html}</div>
    <div class='section'><h3>График тональности по сегментам</h3><canvas id='sentChart' height='180'></canvas></div>

    <script>
      new Chart(document.getElementById('sentChart'), {{
        type: 'bar',
        data: {{
          labels: {json.dumps(chart_labels)},
          datasets: [{{
            data: {json.dumps(chart_scores)},
            backgroundColor: '#3b82f6'
          }}]
        }},
        options: {{
          plugins: {{ legend: {{ display: false }} }},
          scales: {{ y: {{ min: -1, max: 1 }} }}
        }}
      }});
    </script>
    """

    return page(f"Результат — {rec['filename']}", content, "home")


@app.get("/analytics/", response_class=HTMLResponse)
async def analytics_page() -> HTMLResponse:
    records = list(reversed(get_all_records()))

    labels = [str(r.get("meeting_date") or str(r.get("created_at", ""))[:10]) for r in records]
    task_counts = [len(r.get("tasks_json") or []) for r in records]
    qa_counts = [len(r.get("qa_json") or []) for r in records]
    sentiments = [(r.get("analytics_json") or {}).get("avg_sentiment_score", 0) for r in records]
    negative_ratios = [(r.get("analytics_json") or {}).get("negative_ratio", 0) for r in records]

    total_tasks = sum(task_counts)
    total_qa = sum(qa_counts)
    avg_sent = round(sum(sentiments) / len(sentiments), 4) if sentiments else 0.0
    avg_negative = round(sum(negative_ratios) / len(negative_ratios), 4) if negative_ratios else 0.0

    rows = ""

    for r in reversed(records):
        analytics = r.get("analytics_json") or {}
        dynamic = analytics.get("dynamic_analysis") or {}

        rows += f"""
        <tr>
          <td><a href='/result/{r['id']}'>{html.escape(r['filename'])}</a></td>
          <td>{html.escape(r.get('project_name') or '')}</td>
          <td>{r.get('meeting_date') or ''}</td>
          <td>{len(r.get('tasks_json') or [])}</td>
          <td>{len(r.get('qa_json') or [])}</td>
          <td>{analytics.get('avg_sentiment_score', 0)}</td>
          <td>{analytics.get('negative_ratio', 0)}</td>
          <td>{dynamic.get('tasks_delta', 0)}</td>
          <td>{dynamic.get('sentiment_delta', 0)}</td>
        </tr>
        """

    content = f"""
    <div class='cards'>
      <div class='card'><h4>Встреч</h4><div class='val'>{len(records)}</div></div>
      <div class='card'><h4>Всего задач</h4><div class='val'>{total_tasks}</div></div>
      <div class='card'><h4>Всего Q/A</h4><div class='val'>{total_qa}</div></div>
      <div class='card'><h4>Ср. sentiment</h4><div class='val' style='font-size:22px'>{avg_sent}</div></div>
      <div class='card'><h4>Ср. негатив</h4><div class='val' style='font-size:22px'>{avg_negative}</div></div>
    </div>

    <div class='row'>
      <div class='section' style='flex:1'><h3>Динамика тональности</h3><canvas id='sentTrend'></canvas></div>
      <div class='section' style='flex:1'><h3>Задачи по встречам</h3><canvas id='taskTrend'></canvas></div>
    </div>

    <div class='row'>
      <div class='section' style='flex:1'><h3>Q/A по встречам</h3><canvas id='qaTrend'></canvas></div>
      <div class='section' style='flex:1'><h3>Доля негатива</h3><canvas id='negativeTrend'></canvas></div>
    </div>

    <div class='section'>
      <h3>Все встречи</h3>
      <div style='overflow-x:auto'>
        <table>
          <thead>
            <tr>
              <th>Файл</th>
              <th>Проект</th>
              <th>Дата</th>
              <th>Задачи</th>
              <th>Q/A</th>
              <th>Sentiment</th>
              <th>Негатив</th>
              <th>Δ задач</th>
              <th>Δ sentiment</th>
            </tr>
          </thead>
          <tbody>{rows or "<tr><td colspan='9' class='muted'>Нет данных</td></tr>"}</tbody>
        </table>
      </div>
    </div>

    <script>
      new Chart(document.getElementById('sentTrend'), {{
        type: 'line',
        data: {{
          labels: {json.dumps(labels)},
          datasets: [{{
            data: {json.dumps(sentiments)},
            label: 'Sentiment',
            borderColor: '#2563eb',
            fill: false
          }}]
        }},
        options: {{ scales: {{ y: {{ min: -1, max: 1 }} }} }}
      }});

      new Chart(document.getElementById('taskTrend'), {{
        type: 'bar',
        data: {{
          labels: {json.dumps(labels)},
          datasets: [{{
            data: {json.dumps(task_counts)},
            label: 'Tasks',
            backgroundColor: '#8b5cf6'
          }}]
        }},
        options: {{ plugins: {{ legend: {{ display: false }} }} }}
      }});

      new Chart(document.getElementById('qaTrend'), {{
        type: 'bar',
        data: {{
          labels: {json.dumps(labels)},
          datasets: [{{
            data: {json.dumps(qa_counts)},
            label: 'Q/A',
            backgroundColor: '#16a34a'
          }}]
        }},
        options: {{ plugins: {{ legend: {{ display: false }} }} }}
      }});

      new Chart(document.getElementById('negativeTrend'), {{
        type: 'line',
        data: {{
          labels: {json.dumps(labels)},
          datasets: [{{
            data: {json.dumps(negative_ratios)},
            label: 'Negative ratio',
            borderColor: '#dc2626',
            fill: false
          }}]
        }},
        options: {{ scales: {{ y: {{ min: 0, max: 1 }} }} }}
      }});
    </script>
    """

    return page("Аналитика", content, "analytics")


@app.get("/search/", response_class=HTMLResponse)
async def search_page(q: str = "") -> HTMLResponse:
    results = search_records(q) if q else []
    cards = ""

    for r in results:
        cards += f"""
        <div class='task-card'>
          <a href='/result/{r['id']}' style='font-weight:700'>{html.escape(r['filename'])}</a>
          <div class='sub'>{html.escape(r.get('project_name') or '')} | {r.get('meeting_date') or r.get('created_at') or ''}</div>
          <div style='margin-top:6px'>{r.get('snippet') or ''}</div>
        </div>
        """

    content = f"""
    <div class='section'>
      <h3>Поиск по встречам</h3>
      <form method='get' action='/search/'>
        <div class='row'>
          <div style='flex:1'>
            <input type='text' name='q' value='{html.escape(q)}' placeholder='Поиск по транскрипту, проекту, участникам'>
          </div>
          <div><button class='btn btn-blue' type='submit'>Найти</button></div>
        </div>
      </form>
      <div style='margin-top:12px'>
        {cards or ('<p class="muted">Введите запрос</p>' if not q else '<p class="muted">Ничего не найдено</p>')}
      </div>
    </div>
    """

    return page("Поиск", content, "search")


@app.get("/api/analytics", response_class=JSONResponse)
async def api_analytics() -> JSONResponse:
    records = get_all_records()

    payload = {
        "meetings_count": len(records),
        "total_tasks": sum(len(r.get("tasks_json") or []) for r in records),
        "total_qa_pairs": sum(len(r.get("qa_json") or []) for r in records),
        "avg_sentiment_overall": round(
            sum((r.get("analytics_json") or {}).get("avg_sentiment_score", 0) for r in records) / len(records),
            4,
        ) if records else 0,
        "avg_negative_ratio": round(
            sum((r.get("analytics_json") or {}).get("negative_ratio", 0) for r in records) / len(records),
            4,
        ) if records else 0,
    }

    return JSONResponse(payload)


@app.get("/export/excel/{record_id}")
async def excel_export(record_id: int):
    rec = get_record(record_id)

    if not rec:
        raise HTTPException(404, "Record not found")

    path = export_excel(record_id)
    safe_filename = re.sub(r'[^\w\-.]', '_', rec.get("filename", f"record_{record_id}"))

    return FileResponse(
        path,
        filename=f"report_{safe_filename}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/export/pdf/{record_id}")
async def pdf_export(record_id: int):
    rec = get_record(record_id)

    if not rec:
        raise HTTPException(404, "Record not found")

    path = export_pdf(record_id)
    safe_filename = re.sub(r'[^\w\-.]', '_', rec.get("filename", f"record_{record_id}"))

    return FileResponse(
        path,
        filename=f"report_{safe_filename}.pdf",
        media_type="application/pdf",
    )


@app.post("/api/debug/nlp", response_class=JSONResponse)
async def api_debug_nlp(payload: dict) -> JSONResponse:
    """
    Быстрая проверка классификатора без загрузки аудио.
    POST JSON: {"texts": ["Подготовить презентацию к пятнице", ...]}
    """
    texts = payload.get("texts") or []

    if isinstance(texts, str):
        texts = [texts]

    texts = [normalize_text(t) for t in texts if normalize_text(t)]
    classified = final_classify(texts)

    fake_segments = [
        {"start": float(i), "end": float(i + 1), "text": text}
        for i, text in enumerate(texts)
    ]
    fake_sentiments = [{"label": "neutral", "score": 0.0} for _ in texts]
    tasks = detect_tasks(fake_segments, classified, fake_sentiments)

    return JSONResponse({
        "mode": NLP_MODE,
        "classifier_dir": CLASSIFIER_DIR,
        "torch_available": TORCH_AVAILABLE,
        "classifier_loaded": bool(classifier_tokenizer and classifier_model),
        "classified": classified,
        "tasks": tasks,
    })


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "app:app",
        host="127.0.0.1",
        port=8000,
        reload=True,
    )