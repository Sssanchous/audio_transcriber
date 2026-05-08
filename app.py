from __future__ import annotations

import html
import json
import os
import re
import shutil
import time
import hashlib
from pathlib import Path
from typing import Optional

import psycopg
from dotenv import load_dotenv
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydub import AudioSegment
from psycopg.types.json import Json
from faster_whisper import WhisperModel
from task_rule_engine import analyze_task_candidate

TORCH_AVAILABLE = False
NATASHA_AVAILABLE = False
EXCEL_AVAILABLE = False
PDF_AVAILABLE = False

try:
    import torch
    import torch.nn.functional as F
    from transformers import AutoTokenizer, AutoModelForSequenceClassification, pipeline
    TORCH_AVAILABLE = True
except Exception:
    pass

try:
    from natasha import Segmenter, MorphVocab, NewsEmbedding, NewsNERTagger, DatesExtractor, Doc
    NATASHA_AVAILABLE = True
except Exception:
    pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    EXCEL_AVAILABLE = True
except Exception:
    pass

try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
except Exception:
    pass

load_dotenv()
os.environ.setdefault("TOKENIZERS_PARALLELISM", "false")

DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL not found in .env")

WHISPER_MODEL_NAME = os.getenv("WHISPER_MODEL_NAME", "large-v3")
WHISPER_DEVICE = os.getenv("WHISPER_DEVICE", "cpu").lower()
if WHISPER_DEVICE not in {"cpu", "cuda"}:
    WHISPER_DEVICE = "cpu"

WHISPER_COMPUTE_TYPE = os.getenv(
    "WHISPER_COMPUTE_TYPE",
    "float16" if WHISPER_DEVICE == "cuda" else "int8",
)

CLASSIFIER_DIR = os.getenv("CLASSIFIER_DIR", "models/classifier")
SENTIMENT_MODEL_NAME = os.getenv("SENTIMENT_MODEL_NAME", "seara/rubert-tiny2-russian-sentiment")
CLASSIFIER_CONFIDENCE = float(os.getenv("CLASSIFIER_CONFIDENCE_THRESHOLD", "0.78"))
INFERENCE_BATCH_SIZE = int(os.getenv("INFERENCE_BATCH_SIZE", "32"))
QA_MAX_GAP_SECONDS = float(os.getenv("QA_MAX_GAP_SECONDS", "25"))

BASE_DIR = Path(".")
UPLOAD_DIR = BASE_DIR / "uploads"
CONVERTED_DIR = BASE_DIR / "converted"
RESULTS_DIR = BASE_DIR / "results"
EXPORTS_DIR = BASE_DIR / "exports"

for d in [UPLOAD_DIR, CONVERTED_DIR, RESULTS_DIR, EXPORTS_DIR]:
    d.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {".mp3", ".wav", ".m4a", ".opus"}

app = FastAPI(title="PM Insights")
app.mount("/results", StaticFiles(directory=str(RESULTS_DIR)), name="results")
app.mount("/exports", StaticFiles(directory=str(EXPORTS_DIR)), name="exports")

whisper_model = WhisperModel(
    WHISPER_MODEL_NAME,
    device=WHISPER_DEVICE,
    compute_type=WHISPER_COMPUTE_TYPE,
)

classifier_tokenizer = None
classifier_model = None
sentiment_pipe = None
segmenter = morph_vocab = emb_natasha = ner_tagger = dates_extractor = None

NLP_MODE = "rules-only"
LABELS = {0: "task", 1: "question", 2: "answer", 3: "other"}

if TORCH_AVAILABLE:
    cdir = Path(CLASSIFIER_DIR)
    if cdir.exists():
        try:
            classifier_tokenizer = AutoTokenizer.from_pretrained(str(cdir))
            classifier_model = AutoModelForSequenceClassification.from_pretrained(str(cdir))
            device = "cuda" if torch.cuda.is_available() else "cpu"
            classifier_model.to(device)
            classifier_model.eval()
            NLP_MODE = "classifier + rules"
        except Exception as e:
            print(f"[init] classifier disabled: {e}")
            classifier_tokenizer = None
            classifier_model = None

    try:
        sentiment_pipe = pipeline(
            "text-classification",
            model=SENTIMENT_MODEL_NAME,
            tokenizer=SENTIMENT_MODEL_NAME,
            device=0 if torch.cuda.is_available() else -1,
            truncation=True,
        )
        NLP_MODE += " + sentiment"
    except Exception as e:
        print(f"[init] sentiment disabled: {e}")
        sentiment_pipe = None

if NATASHA_AVAILABLE:
    try:
        segmenter = Segmenter()
        morph_vocab = MorphVocab()
        emb_natasha = NewsEmbedding()
        ner_tagger = NewsNERTagger(emb_natasha)
        dates_extractor = DatesExtractor(morph_vocab)
        NLP_MODE += " + natasha"
    except Exception as e:
        print(f"[init] natasha disabled: {e}")
        segmenter = morph_vocab = emb_natasha = ner_tagger = dates_extractor = None

FILLER_WORDS = {
    "ну", "вот", "как", "бы", "типа", "короче", "значит", "получается", "наверное", "наверно",
    "вообще", "просто", "собственно", "то", "есть", "там", "тут", "это", "этот", "эта", "эти",
    "какой-то", "какая-то", "какие-то", "что-то", "как-то", "где-то", "может", "может быть",
}

INTERROGATIVE_WORDS = (
    "кто", "что", "где", "когда", "почему", "зачем", "как", "какой", "какая", "какие", "сколько",
    "можно ли", "нужно ли", "успеем ли", "в чём", "а что", "что за",
)

ANSWER_STARTERS = [
    r"^да\b", r"^нет\b", r"^хорошо\b", r"^ладно\b", r"^ок(ей)?\b", r"^понял\b", r"^согласен\b",
    r"^согласна\b", r"^верно\b", r"^именно\b", r"^конечно\b", r"^готово\b", r"^принято\b",
]

ANSWER_PATTERNS = [
    r"\bя\s+(сделаю|подготовлю|отправлю|проверю|исправлю|обновлю|настрою|доделаю|согласую)\b",
    r"\bмы\s+(сделаем|подготовим|отправим|проверим|исправим|обновим)\b",
    r"\b(сделано|готово|выполнено|завершено|загружено|отправлено)\b",
    r"\bза\s+это\s+отвечает\b",
    r"\bуже\s+(готов|готова|сделал|сделала|отправил|отправила)\b",
]

PROGRESS_PATTERNS = [
    r"\bесть\s+умерен(ный|ное|ная)\s+прогресс\b",
    r"\bмы\s+уже\s+синхронизировались\b",
    r"\bсейчас\s+коротко\s+расскажу\b",
    r"\bна\s+прошлой\s+неделе\s+мы\s+обсуждали\b",
    r"\bслово\s+передаю\b",
    r"\bобновление\s+по\s+своему\s+блоку\b",
    r"\bпрогресс\s+хороший\b",
]

SPEAKER_PREFIX_RE = re.compile(r"^\s*(Говорящий|Спикер|Speaker)\s*\d*\s*:\s*", re.I)
PERSON_PREFIX_RE = re.compile(r"^\s*([А-ЯЁ][а-яё]+)\s*,\s*(.+)$")

MIN_QA_CHARS = 18


def normalize_text(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text.strip(" -–—")


def strip_speaker_prefix(text: str) -> str:
    text = SPEAKER_PREFIX_RE.sub("", text)
    return normalize_text(text)


def is_filler_heavy(text: str) -> bool:
    words = [
        w.strip(".,!?;:()[]\"'«»").lower()
        for w in text.split()
        if w.strip(".,!?;:()[]\"'«»")
    ]
    if len(words) <= 3:
        return False
    filler_count = sum(1 for w in words if w in FILLER_WORDS)
    return filler_count / max(len(words), 1) >= 0.45


def is_progress_text(text: str) -> bool:
    lower = text.lower()
    return any(re.search(p, lower) for p in PROGRESS_PATTERNS)


def extract_deadline(text: str) -> Optional[str]:
    analysis = analyze_task_candidate(text)
    reasons = analysis.get("reasons", [])
    for r in reasons:
        if r == "deadline":
            lower = text.lower()
            patterns = [
                r"\bдо\s+\d{1,2}[:.]\d{2}\b",
                r"\bдо\s+\d{1,2}[./]\d{1,2}(?:[./]\d{2,4})?\b",
                r"\bдо\s+(понедельника|вторника|среды|четверга|пятницы|субботы|воскресенья)\b",
                r"\bдо\s+(завтра|вечера|утра|обеда|релиза|конца\s+дня|конца\s+недели|следующей\s+встречи)\b",
                r"\bк\s+(утру|вечеру|обеду|созвону|завтрашнему\s+созвону)\b",
                r"\bсегодня\b",
                r"\bзавтра\b",
                r"\bна\s+этой\s+неделе\b",
                r"\bпосле\s+обеда\b",
            ]
            for p in patterns:
                m = re.search(p, lower)
                if m:
                    return m.group(0)
    return None


def extract_responsible(text: str) -> Optional[str]:
    m = PERSON_PREFIX_RE.match(text)
    if m:
        return m.group(1)

    if NATASHA_AVAILABLE and ner_tagger and segmenter:
        try:
            doc = Doc(text)
            doc.segment(segmenter)
            doc.tag_ner(ner_tagger)
            for span in doc.spans:
                if span.type == "PER":
                    span.normalize(morph_vocab)
                    return span.normal if getattr(span, "normal", None) else span.text
        except Exception:
            pass

    m = re.search(r"\bза\s+это\s+отвечает\s+([А-ЯЁ][а-яё]+)\b", text)
    if m:
        return m.group(1)

    return None


def classify_by_rules(text: str) -> tuple[str, float, dict]:
    text = strip_speaker_prefix(text)
    lower = text.lower()

    debug = {
        "filler": is_filler_heavy(text),
        "progress": is_progress_text(text),
    }

    if len(text) < MIN_QA_CHARS or is_filler_heavy(text):
        return "other", 0.95, debug

    if text.endswith("?") or any(lower.startswith(w + " ") for w in INTERROGATIVE_WORDS):
        return "question", 0.97, debug

    if any(re.search(p, lower) for p in ANSWER_STARTERS) or any(re.search(p, lower) for p in ANSWER_PATTERNS):
        return "answer", 0.93, debug

    if is_progress_text(text):
        return "other", 0.97, debug

    task_analysis = analyze_task_candidate(text)
    debug["task_rule_engine"] = task_analysis

    if task_analysis["is_task"]:
        return "task", max(0.9, float(task_analysis["score"])), debug

    return "other", 0.9, debug


def batch_classifier_predict(texts: list[str]) -> list[tuple[str, float]]:
    if not (TORCH_AVAILABLE and classifier_tokenizer and classifier_model):
        return [("other", 0.0) for _ in texts]

    device = next(classifier_model.parameters()).device
    out: list[tuple[str, float]] = []

    for i in range(0, len(texts), INFERENCE_BATCH_SIZE):
        batch = texts[i:i + INFERENCE_BATCH_SIZE]
        enc = classifier_tokenizer(
            batch,
            truncation=True,
            padding=True,
            max_length=128,
            return_tensors="pt",
        )
        enc = {k: v.to(device) for k, v in enc.items()}

        with torch.no_grad():
            logits = classifier_model(**enc).logits
            probs = F.softmax(logits, dim=-1)
            confs, idxs = probs.max(dim=-1)

        for conf, idx in zip(confs.tolist(), idxs.tolist()):
            out.append((LABELS.get(idx, "other"), float(conf)))

    return out


def final_classify(texts: list[str]) -> list[dict]:
    rule_results = [classify_by_rules(t) for t in texts]
    clf_results = batch_classifier_predict(texts)
    final: list[dict] = []

    for text, (rule_label, rule_conf, debug), (clf_label, clf_conf) in zip(texts, rule_results, clf_results):
        label = rule_label
        confidence = rule_conf
        source = "rules"

        task_analysis = analyze_task_candidate(text)

        if clf_conf >= CLASSIFIER_CONFIDENCE and rule_label != "other":
            if clf_label == rule_label:
                confidence = max(rule_conf, clf_conf)
                source = "rules+classifier"
        elif clf_conf >= max(CLASSIFIER_CONFIDENCE + 0.08, 0.86) and rule_label == "other":
            if not debug.get("filler") and not is_progress_text(text):
                if clf_label in {"question", "answer"}:
                    label = clf_label
                    confidence = clf_conf
                    source = "classifier"
                elif clf_label == "task" and task_analysis["is_task"]:
                    label = "task"
                    confidence = max(clf_conf, float(task_analysis["score"]))
                    source = "classifier+task_rule_engine"

        if label == "task" and not task_analysis["is_task"]:
            label = "other"
            confidence = 0.49
            source = "task_rule_engine_reject"

        final.append({
            "text": text,
            "label": label,
            "confidence": round(float(confidence), 4),
            "source": source,
            "debug": debug,
            "task_debug": task_analysis,
        })

    return final


def batch_sentiment(texts: list[str]) -> list[dict]:
    if not sentiment_pipe:
        return [{"label": "neutral", "score": 0.0} for _ in texts]

    results: list[dict] = []

    for i in range(0, len(texts), INFERENCE_BATCH_SIZE):
        batch = texts[i:i + INFERENCE_BATCH_SIZE]

        try:
            preds = sentiment_pipe(batch)
        except Exception:
            preds = [{"label": "neutral", "score": 0.0} for _ in batch]

        for pred in preds:
            label = str(pred.get("label", "neutral")).lower()
            score = float(pred.get("score", 0.0))

            if "neg" in label:
                results.append({"label": "negative", "score": round(-score, 4)})
            elif "pos" in label:
                results.append({"label": "positive", "score": round(score, 4)})
            else:
                results.append({"label": "neutral", "score": 0.0})

    return results


def is_good_question(text: str) -> bool:
    text = strip_speaker_prefix(text)

    if len(text) < MIN_QA_CHARS:
        return False

    if is_filler_heavy(text):
        return False

    lower = text.lower()
    return text.endswith("?") or any(lower.startswith(w + " ") for w in INTERROGATIVE_WORDS)


def is_good_answer(text: str) -> bool:
    text = strip_speaker_prefix(text)

    if len(text) < MIN_QA_CHARS:
        return False

    if is_filler_heavy(text):
        return False

    lower = text.lower()

    if any(re.search(p, lower) for p in ANSWER_STARTERS):
        return True

    if any(re.search(p, lower) for p in ANSWER_PATTERNS):
        return True

    if is_progress_text(text):
        return False

    return len(text.split()) >= 4


def build_qa_pairs(segments: list[dict], classified: list[dict]) -> list[dict]:
    pairs: list[dict] = []
    pending_questions: list[dict] = []

    for seg, cls in zip(segments, classified):
        text = seg["text"]

        if cls["label"] == "question" and is_good_question(text):
            pending_questions.append({"seg": seg, "cls": cls})
            continue

        if cls["label"] == "answer" and is_good_answer(text):
            best_q = None

            for q in reversed(pending_questions):
                gap = seg["start"] - q["seg"]["end"]

                if gap < 0:
                    continue

                if gap > QA_MAX_GAP_SECONDS:
                    continue

                best_q = q
                break

            if best_q:
                qseg = best_q["seg"]
                conf = round((best_q["cls"]["confidence"] + cls["confidence"]) / 2, 4)

                pairs.append({
                    "question": qseg["text"],
                    "answer": seg["text"],
                    "question_timecode": f"{qseg['start']:.2f} - {qseg['end']:.2f} сек.",
                    "answer_timecode": f"{seg['start']:.2f} - {seg['end']:.2f} сек.",
                    "confidence": conf,
                })

                pending_questions.remove(best_q)

    return pairs


def dedupe_tasks(tasks: list[dict]) -> list[dict]:
    seen: set[str] = set()
    result: list[dict] = []

    for task in tasks:
        key = re.sub(r"\s+", " ", task["text"].lower()).strip(" .,!?")

        if key in seen:
            continue

        seen.add(key)
        result.append(task)

    return result


def normalize_task_text(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[^\w\sа-яё]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    stop_words = {
        "нужно", "надо", "сделать", "подготовить", "проверить", "доделать",
        "задача", "важно", "давайте", "пожалуйста", "это", "тот", "та", "те",
        "к", "до", "на", "по", "и", "или", "в", "во", "с", "со", "для",
        "будет", "будем", "можно", "еще", "ещё", "там", "тут", "вот",
    }

    words = [w for w in text.split() if w not in stop_words and len(w) > 2]
    return " ".join(words)


def text_similarity(a: str, b: str) -> float:
    a_words = set(normalize_task_text(a).split())
    b_words = set(normalize_task_text(b).split())

    if not a_words or not b_words:
        return 0.0

    intersection = len(a_words & b_words)
    union = len(a_words | b_words)

    return round(intersection / union, 4) if union else 0.0


def compare_tasks_with_previous(current_tasks: list[dict], previous_tasks: list[dict]) -> dict:
    repeated_tasks = []
    new_tasks = []
    potentially_closed_tasks = []
    matched_previous_indexes = set()

    for current_task in current_tasks:
        best_match = None
        best_score = 0.0
        best_index = None

        for idx, previous_task in enumerate(previous_tasks):
            score = text_similarity(
                current_task.get("text", ""),
                previous_task.get("text", ""),
            )

            if score > best_score:
                best_score = score
                best_match = previous_task
                best_index = idx

        if best_match and best_score >= 0.45:
            repeated_tasks.append({
                "current_task": current_task,
                "previous_task": best_match,
                "similarity": best_score,
            })

            if best_index is not None:
                matched_previous_indexes.add(best_index)
        else:
            new_tasks.append(current_task)

    for idx, previous_task in enumerate(previous_tasks):
        if idx not in matched_previous_indexes:
            potentially_closed_tasks.append(previous_task)

    return {
        "new_tasks": new_tasks,
        "repeated_tasks": repeated_tasks,
        "potentially_closed_tasks": potentially_closed_tasks,
        "new_count": len(new_tasks),
        "repeated_count": len(repeated_tasks),
        "potentially_closed_count": len(potentially_closed_tasks),
    }


def detect_tasks(segments: list[dict], classified: list[dict], sentiments: list[dict]) -> list[dict]:
    tasks: list[dict] = []

    for seg, cls, sent in zip(segments, classified, sentiments):
        if cls["label"] != "task":
            continue

        text = seg["text"]
        task_analysis = analyze_task_candidate(text)

        if not task_analysis["is_task"]:
            continue

        responsible = extract_responsible(text)
        deadline = extract_deadline(text)

        tasks.append({
            "text": text,
            "responsible": responsible,
            "deadline": deadline,
            "confidence": round(max(float(cls["confidence"]), float(task_analysis["score"])), 4),
            "sentiment": sent,
            "status": "new",
            "timecode": f"{seg['start']:.2f} - {seg['end']:.2f} сек.",
            "task_debug": task_analysis,
        })

    return dedupe_tasks(tasks)


def detect_problem_aspects(text: str) -> list[str]:
    patterns = {
        "сроки": [r"\bсрок\b", r"\bдедлайн\b", r"\bзадерж", r"\bдо\s+пятницы\b"],
        "ошибки": [r"\bошиб", r"\bбаг\b", r"\bне\s+работает\b", r"\bсломано\b"],
        "сервер": [r"\bсервер\b", r"\bдеплой\b", r"\bрелиз\b", r"\bинфраструктур"],
        "документы": [r"\bотч[её]т\b", r"\bдоговор\b", r"\bдокумент\b", r"\bпрезентац"],
        "ресурсы": [r"\bресурс", r"\bне\s+успеваем\b", r"\bперегруз"],
    }

    lower = text.lower()
    found = []

    for name, pats in patterns.items():
        if any(re.search(p, lower) for p in pats):
            found.append(name)

    return found


def get_conn():
    return psycopg.connect(DATABASE_URL)


def init_db() -> None:
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS transcriptions (
                id SERIAL PRIMARY KEY,
                filename TEXT NOT NULL,
                original_path TEXT,
                converted_path TEXT,
                duration DOUBLE PRECISION DEFAULT 0,
                language TEXT DEFAULT 'unknown',
                full_text TEXT DEFAULT '',
                segments JSONB DEFAULT '[]',
                tasks_json JSONB DEFAULT '[]',
                qa_json JSONB DEFAULT '[]',
                sentiment_json JSONB DEFAULT '[]',
                analytics_json JSONB DEFAULT '{}',
                meeting_date DATE,
                project_name TEXT DEFAULT '',
                participants TEXT DEFAULT '',
                timing_json JSONB DEFAULT '{}',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                search_vector tsvector
            )
            """
        )
        cur.execute("CREATE INDEX IF NOT EXISTS idx_transcriptions_search ON transcriptions USING GIN(search_vector)")
        conn.commit()


def update_search_vector(cur, record_id: int) -> None:
    cur.execute(
        """
        UPDATE transcriptions
        SET search_vector =
            to_tsvector(
                'russian',
                coalesce(filename,'') || ' ' ||
                coalesce(project_name,'') || ' ' ||
                coalesce(participants,'') || ' ' ||
                coalesce(full_text,'')
            )
        WHERE id = %s
        """,
        (record_id,),
    )


def get_record(record_id: int) -> Optional[dict]:
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT * FROM transcriptions WHERE id = %s", (record_id,))
        row = cur.fetchone()

        if not row:
            return None

        cols = [d[0] for d in cur.description]
        return dict(zip(cols, row))


def get_all_records() -> list[dict]:
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute("SELECT * FROM transcriptions ORDER BY created_at DESC")
        rows = cur.fetchall()
        cols = [d[0] for d in cur.description]

        return [dict(zip(cols, row)) for row in rows]


def find_previous_record(current_record_id: int, project_name: str = "") -> Optional[dict]:
    with get_conn() as conn, conn.cursor() as cur:
        if project_name:
            cur.execute(
                """
                SELECT *
                FROM transcriptions
                WHERE id < %s AND project_name = %s
                ORDER BY meeting_date DESC NULLS LAST, created_at DESC
                LIMIT 1
                """,
                (current_record_id, project_name),
            )
        else:
            cur.execute(
                """
                SELECT *
                FROM transcriptions
                WHERE id < %s
                ORDER BY meeting_date DESC NULLS LAST, created_at DESC
                LIMIT 1
                """,
                (current_record_id,),
            )

        row = cur.fetchone()

        if not row:
            return None

        cols = [d[0] for d in cur.description]
        return dict(zip(cols, row))


def search_records(query: str) -> list[dict]:
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            SELECT id, filename, project_name, meeting_date, created_at,
                   ts_headline(
                       'russian',
                       full_text,
                       plainto_tsquery('russian', %s),
                       'MaxWords=30, MinWords=10, StartSel=<mark>, StopSel=</mark>'
                   ) AS snippet,
                   ts_rank(search_vector, plainto_tsquery('russian', %s)) AS rank
            FROM transcriptions
            WHERE search_vector @@ plainto_tsquery('russian', %s)
            ORDER BY rank DESC, created_at DESC
            LIMIT 30
            """,
            (query, query, query),
        )

        return [
            {
                "id": r[0],
                "filename": r[1],
                "project_name": r[2],
                "meeting_date": r[3],
                "created_at": r[4],
                "snippet": r[5],
                "rank": float(r[6]),
            }
            for r in cur.fetchall()
        ]


def convert_to_wav(input_path: Path, output_path: Path) -> None:
    audio = AudioSegment.from_file(input_path)
    audio = audio.set_channels(1).set_frame_rate(16000)
    audio.export(output_path, format="wav")


def transcribe_audio(wav_path: Path) -> dict:
    segments_iter, info = whisper_model.transcribe(
        str(wav_path),
        beam_size=5,
        vad_filter=True,
    )

    segments = []
    text_parts = []

    for seg in segments_iter:
        text = normalize_text(seg.text)

        if not text:
            continue

        segments.append({
            "start": round(seg.start, 2),
            "end": round(seg.end, 2),
            "text": text,
        })

        text_parts.append(text)

    return {
        "language": getattr(info, "language", "unknown") or "unknown",
        "duration": round(getattr(info, "duration", 0.0) or 0.0, 2),
        "segments": segments,
        "text": " ".join(text_parts),
    }


def process_segments(segments: list[dict]) -> tuple[list[dict], list[dict], list[dict], list[dict], dict]:
    texts = [strip_speaker_prefix(s["text"]) for s in segments]
    classified = final_classify(texts)
    sentiments = batch_sentiment(texts)
    tasks = detect_tasks(segments, classified, sentiments)
    qa_pairs = build_qa_pairs(segments, classified)

    sent_values = [s["score"] for s in sentiments]
    avg_sent = round(sum(sent_values) / len(sent_values), 4) if sent_values else 0.0
    neg_ratio = round(
        sum(1 for s in sentiments if s["label"] == "negative") / len(sentiments),
        4,
    ) if sentiments else 0.0

    aspect_counts: dict[str, int] = {}

    for seg in segments:
        for asp in detect_problem_aspects(seg["text"]):
            aspect_counts[asp] = aspect_counts.get(asp, 0) + 1

    analytics = {
        "tasks_count": len(tasks),
        "qa_count": len(qa_pairs),
        "avg_sentiment_score": avg_sent,
        "negative_ratio": neg_ratio,
        "top_problem_aspects": sorted(
            [{"aspect": k, "count": v} for k, v in aspect_counts.items()],
            key=lambda x: -x["count"],
        )[:5],
    }

    return classified, sentiments, tasks, qa_pairs, analytics


def build_dynamic_analysis(current_record: dict, previous_record: Optional[dict]) -> dict:
    current_tasks = current_record.get("tasks_json") or []
    current_qa = current_record.get("qa_json") or []
    current_analytics = current_record.get("analytics_json") or {}

    if not previous_record:
        return {
            "has_previous": False,
            "summary": "Предыдущая встреча для сравнения не найдена.",
            "tasks_delta": 0,
            "qa_delta": 0,
            "sentiment_delta": 0,
            "negative_ratio_delta": 0,
            "task_changes": {
                "new_tasks": current_tasks,
                "repeated_tasks": [],
                "potentially_closed_tasks": [],
                "new_count": len(current_tasks),
                "repeated_count": 0,
                "potentially_closed_count": 0,
            },
            "aspect_changes": [],
        }

    previous_tasks = previous_record.get("tasks_json") or []
    previous_qa = previous_record.get("qa_json") or []
    previous_analytics = previous_record.get("analytics_json") or {}

    current_sentiment = float(current_analytics.get("avg_sentiment_score", 0) or 0)
    previous_sentiment = float(previous_analytics.get("avg_sentiment_score", 0) or 0)

    current_negative = float(current_analytics.get("negative_ratio", 0) or 0)
    previous_negative = float(previous_analytics.get("negative_ratio", 0) or 0)

    task_changes = compare_tasks_with_previous(current_tasks, previous_tasks)

    current_aspects = {
        item.get("aspect"): item.get("count", 0)
        for item in current_analytics.get("top_problem_aspects", [])
        if item.get("aspect")
    }

    previous_aspects = {
        item.get("aspect"): item.get("count", 0)
        for item in previous_analytics.get("top_problem_aspects", [])
        if item.get("aspect")
    }

    aspect_changes = []

    for aspect in sorted(set(current_aspects) | set(previous_aspects)):
        current_count = current_aspects.get(aspect, 0)
        previous_count = previous_aspects.get(aspect, 0)

        aspect_changes.append({
            "aspect": aspect,
            "current_count": current_count,
            "previous_count": previous_count,
            "delta": current_count - previous_count,
        })

    tasks_delta = len(current_tasks) - len(previous_tasks)
    qa_delta = len(current_qa) - len(previous_qa)
    sentiment_delta = round(current_sentiment - previous_sentiment, 4)
    negative_ratio_delta = round(current_negative - previous_negative, 4)

    summary_parts = []

    if tasks_delta > 0:
        summary_parts.append(f"Количество задач увеличилось на {tasks_delta}.")
    elif tasks_delta < 0:
        summary_parts.append(f"Количество задач уменьшилось на {abs(tasks_delta)}.")
    else:
        summary_parts.append("Количество задач не изменилось.")

    if qa_delta > 0:
        summary_parts.append(f"Количество пар «вопрос — ответ» увеличилось на {qa_delta}.")
    elif qa_delta < 0:
        summary_parts.append(f"Количество пар «вопрос — ответ» уменьшилось на {abs(qa_delta)}.")
    else:
        summary_parts.append("Количество пар «вопрос — ответ» не изменилось.")

    if sentiment_delta > 0:
        summary_parts.append("Средняя эмоциональная окраска стала более позитивной.")
    elif sentiment_delta < 0:
        summary_parts.append("Средняя эмоциональная окраска стала более негативной.")
    else:
        summary_parts.append("Средняя эмоциональная окраска не изменилась.")

    if negative_ratio_delta > 0:
        summary_parts.append("Доля негативных сегментов выросла.")
    elif negative_ratio_delta < 0:
        summary_parts.append("Доля негативных сегментов снизилась.")
    else:
        summary_parts.append("Доля негативных сегментов осталась прежней.")

    return {
        "has_previous": True,
        "previous_record_id": previous_record.get("id"),
        "previous_filename": previous_record.get("filename"),
        "previous_meeting_date": str(previous_record.get("meeting_date") or ""),
        "tasks_delta": tasks_delta,
        "qa_delta": qa_delta,
        "sentiment_delta": sentiment_delta,
        "negative_ratio_delta": negative_ratio_delta,
        "task_changes": task_changes,
        "aspect_changes": aspect_changes,
        "summary": " ".join(summary_parts),
    }


def save_result_files(stem: str, payload: dict) -> tuple[Path, Path]:
    json_path = RESULTS_DIR / f"{stem}.json"
    txt_path = RESULTS_DIR / f"{stem}.txt"

    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    txt_path.write_text(payload.get("full_text", ""), encoding="utf-8")

    return json_path, txt_path


def store_record(
    filename: str,
    original_path: str,
    converted_path: str,
    transcription: dict,
    tasks: list[dict],
    qa_pairs: list[dict],
    sentiments: list[dict],
    analytics: dict,
    meeting_date: Optional[str],
    project_name: str,
    participants: str,
    timing: dict,
) -> int:
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO transcriptions (
                filename,
                original_path,
                converted_path,
                duration,
                language,
                full_text,
                segments,
                tasks_json,
                qa_json,
                sentiment_json,
                analytics_json,
                meeting_date,
                project_name,
                participants,
                timing_json
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (
                filename,
                original_path,
                converted_path,
                transcription["duration"],
                transcription["language"],
                transcription["text"],
                Json(transcription["segments"]),
                Json(tasks),
                Json(qa_pairs),
                Json(sentiments),
                Json(analytics),
                meeting_date or None,
                project_name,
                participants,
                Json(timing),
            ),
        )

        rid = cur.fetchone()[0]
        update_search_vector(cur, rid)
        conn.commit()

        return rid


def export_excel(record_id: int) -> Path:
    if not EXCEL_AVAILABLE:
        raise HTTPException(500, "openpyxl not installed")

    rec = get_record(record_id)

    if not rec:
        raise HTTPException(404, "Record not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    bold = Font(bold=True)

    analytics = rec.get("analytics_json") or {}

    rows = [
        ("Файл", rec["filename"]),
        ("Проект", rec.get("project_name") or ""),
        ("Дата", str(rec.get("meeting_date") or "")),
        ("Участники", rec.get("participants") or ""),
        ("Длительность, сек", rec.get("duration") or 0),
        ("Средний sentiment", analytics.get("avg_sentiment_score", 0)),
        ("Негативная доля", analytics.get("negative_ratio", 0)),
    ]

    for i, (k, v) in enumerate(rows, 1):
        ws.cell(i, 1, k).font = bold
        ws.cell(i, 2, v)

    tasks_ws = wb.create_sheet("Задачи")
    task_headers = ["Текст", "Ответственный", "Срок", "Confidence", "Sentiment", "Timecode"]

    for i, h in enumerate(task_headers, 1):
        tasks_ws.cell(1, i, h).font = bold

    for r, task in enumerate(rec.get("tasks_json") or [], 2):
        tasks_ws.cell(r, 1, task.get("text", ""))
        tasks_ws.cell(r, 2, task.get("responsible", ""))
        tasks_ws.cell(r, 3, task.get("deadline", ""))
        tasks_ws.cell(r, 4, task.get("confidence", 0))
        tasks_ws.cell(r, 5, (task.get("sentiment") or {}).get("label", "neutral"))
        tasks_ws.cell(r, 6, task.get("timecode", ""))

    qa_ws = wb.create_sheet("QA")
    qa_headers = ["Вопрос", "Ответ", "Question time", "Answer time", "Confidence"]

    for i, h in enumerate(qa_headers, 1):
        qa_ws.cell(1, i, h).font = bold

    for r, qa in enumerate(rec.get("qa_json") or [], 2):
        qa_ws.cell(r, 1, qa.get("question", ""))
        qa_ws.cell(r, 2, qa.get("answer", ""))
        qa_ws.cell(r, 3, qa.get("question_timecode", ""))
        qa_ws.cell(r, 4, qa.get("answer_timecode", ""))
        qa_ws.cell(r, 5, qa.get("confidence", 0))

    dynamic = analytics.get("dynamic_analysis") or {}
    dyn_ws = wb.create_sheet("Динамика")

    dyn_rows = [
        ("Есть предыдущая встреча", "Да" if dynamic.get("has_previous") else "Нет"),
        ("Предыдущий файл", dynamic.get("previous_filename", "")),
        ("Дата предыдущей встречи", dynamic.get("previous_meeting_date", "")),
        ("Изменение количества задач", dynamic.get("tasks_delta", 0)),
        ("Изменение количества Q/A", dynamic.get("qa_delta", 0)),
        ("Изменение sentiment", dynamic.get("sentiment_delta", 0)),
        ("Изменение негативной доли", dynamic.get("negative_ratio_delta", 0)),
        ("Вывод", dynamic.get("summary", "")),
    ]

    for i, (k, v) in enumerate(dyn_rows, 1):
        dyn_ws.cell(i, 1, k).font = bold
        dyn_ws.cell(i, 2, v)

    task_changes = dynamic.get("task_changes") or {}

    dyn_ws.cell(10, 1, "Новые задачи").font = bold
    row_idx = 11

    for task in task_changes.get("new_tasks", []):
        dyn_ws.cell(row_idx, 1, task.get("text", ""))
        dyn_ws.cell(row_idx, 2, task.get("responsible", ""))
        dyn_ws.cell(row_idx, 3, task.get("deadline", ""))
        row_idx += 1

    row_idx += 1
    dyn_ws.cell(row_idx, 1, "Повторяющиеся задачи").font = bold
    row_idx += 1

    for item in task_changes.get("repeated_tasks", []):
        dyn_ws.cell(row_idx, 1, (item.get("current_task") or {}).get("text", ""))
        dyn_ws.cell(row_idx, 2, (item.get("previous_task") or {}).get("text", ""))
        dyn_ws.cell(row_idx, 3, item.get("similarity", 0))
        row_idx += 1

    row_idx += 1
    dyn_ws.cell(row_idx, 1, "Потенциально закрытые задачи").font = bold
    row_idx += 1

    for task in task_changes.get("potentially_closed_tasks", []):
        dyn_ws.cell(row_idx, 1, task.get("text", ""))
        dyn_ws.cell(row_idx, 2, task.get("responsible", ""))
        dyn_ws.cell(row_idx, 3, task.get("deadline", ""))
        row_idx += 1

    row_idx += 1
    dyn_ws.cell(row_idx, 1, "Изменение проблемных аспектов").font = bold
    row_idx += 1

    dyn_ws.cell(row_idx, 1, "Аспект").font = bold
    dyn_ws.cell(row_idx, 2, "Прошлая встреча").font = bold
    dyn_ws.cell(row_idx, 3, "Текущая встреча").font = bold
    dyn_ws.cell(row_idx, 4, "Изменение").font = bold
    row_idx += 1

    for item in dynamic.get("aspect_changes", []):
        dyn_ws.cell(row_idx, 1, item.get("aspect", ""))
        dyn_ws.cell(row_idx, 2, item.get("previous_count", 0))
        dyn_ws.cell(row_idx, 3, item.get("current_count", 0))
        dyn_ws.cell(row_idx, 4, item.get("delta", 0))
        row_idx += 1

    out = EXPORTS_DIR / f"report_{record_id}.xlsx"
    wb.save(out)

    return out


def export_pdf(record_id: int) -> Path:
    if not PDF_AVAILABLE:
        raise HTTPException(500, "fpdf not installed")

    rec = get_record(record_id)

    if not rec:
        raise HTTPException(404, "Record not found")

    class PDF(FPDF):
        pass

    pdf = PDF()
    pdf.add_page()

    font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"

    if Path(font_path).exists():
        pdf.add_font("DejaVu", "", font_path)
        pdf.set_font("DejaVu", size=12)
    else:
        pdf.set_font("Arial", size=12)

    analytics = rec.get("analytics_json") or {}
    dynamic = analytics.get("dynamic_analysis") or {}

    pdf.multi_cell(0, 8, f"PM Insights — {rec['filename']}")
    pdf.ln(2)
    pdf.multi_cell(0, 8, f"Проект: {rec.get('project_name') or ''}")
    pdf.multi_cell(0, 8, f"Дата: {rec.get('meeting_date') or ''}")
    pdf.multi_cell(0, 8, f"Участники: {rec.get('participants') or ''}")
    pdf.multi_cell(0, 8, f"Длительность: {rec.get('duration') or 0} сек")
    pdf.ln(2)

    pdf.multi_cell(0, 8, "Динамический анализ:")
    pdf.multi_cell(0, 8, f"Вывод: {dynamic.get('summary', 'Нет данных')}")
    pdf.multi_cell(0, 8, f"Изменение количества задач: {dynamic.get('tasks_delta', 0)}")
    pdf.multi_cell(0, 8, f"Изменение количества Q/A: {dynamic.get('qa_delta', 0)}")
    pdf.multi_cell(0, 8, f"Изменение sentiment: {dynamic.get('sentiment_delta', 0)}")
    pdf.multi_cell(0, 8, f"Изменение негативной доли: {dynamic.get('negative_ratio_delta', 0)}")
    pdf.ln(2)

    pdf.multi_cell(0, 8, "Задачи:")

    for i, task in enumerate(rec.get("tasks_json") or [], 1):
        line = (
            f"{i}. {task.get('text', '')} | "
            f"Ответственный: {task.get('responsible') or '—'} | "
            f"Срок: {task.get('deadline') or '—'}"
        )
        pdf.multi_cell(0, 8, line)

    out = EXPORTS_DIR / f"report_{record_id}.pdf"
    pdf.output(str(out))

    return out


def page(title: str, content: str, active: str = "home") -> HTMLResponse:
    nav = f"""
    <nav class='nav'>
      <a class='{ 'active' if active == 'home' else '' }' href='/'>Загрузка</a>
      <a class='{ 'active' if active == 'analytics' else '' }' href='/analytics/'>Аналитика</a>
      <a class='{ 'active' if active == 'search' else '' }' href='/search/'>Поиск</a>
    </nav>
    """

    html_doc = f"""
    <!doctype html>
    <html lang='ru'>
    <head>
      <meta charset='utf-8'>
      <meta name='viewport' content='width=device-width, initial-scale=1'>
      <title>{html.escape(title)}</title>
      <script src='https://cdn.jsdelivr.net/npm/chart.js'></script>
      <style>
        body {{ font-family: Arial, sans-serif; margin: 0; background: #f5f7fb; color: #1f2937; }}
        .wrap {{ max-width: 1180px; margin: 0 auto; padding: 20px; }}
        .nav {{ display:flex; gap:16px; padding:14px 20px; background:#111827; }}
        .nav a {{ color:#cbd5e1; text-decoration:none; font-weight:600; }}
        .nav a.active {{ color:#fff; }}
        .cards {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(190px,1fr)); gap:14px; margin:16px 0; }}
        .card, .section {{ background:#fff; border-radius:12px; padding:16px; box-shadow:0 2px 8px rgba(0,0,0,.05); }}
        .card h4 {{ margin:0 0 8px; color:#6b7280; font-size:13px; }}
        .val {{ font-size:28px; font-weight:700; }}
        .sub {{ font-size:12px; color:#6b7280; margin-top:6px; }}
        .btn {{ display:inline-block; padding:10px 14px; border-radius:8px; text-decoration:none; color:#fff; border:none; cursor:pointer; }}
        .btn-blue {{ background:#2563eb; }}
        .btn-green {{ background:#16a34a; }}
        .btn-red {{ background:#dc2626; }}
        .btn-gray {{ background:#6b7280; }}
        .row {{ display:flex; gap:10px; flex-wrap:wrap; }}
        input[type=text], input[type=date], textarea {{ width:100%; padding:10px 12px; border:1px solid #d1d5db; border-radius:8px; box-sizing:border-box; }}
        label {{ font-weight:600; font-size:14px; display:block; margin-bottom:6px; }}
        .task-card {{ border:1px solid #e5e7eb; border-radius:10px; padding:12px; margin:10px 0; background:#fff; }}
        .seg {{ padding:10px 12px; border-left:6px solid #d1d5db; background:#f9fafb; margin:8px 0; border-radius:8px; }}
        .seg-pos {{ border-color:#10b981; }}
        .seg-neg {{ border-color:#ef4444; }}
        .seg-neu {{ border-color:#9ca3af; }}
        .timing {{ font-size:14px; color:#4b5563; display:flex; gap:16px; flex-wrap:wrap; margin:12px 0; }}
        table {{ width:100%; border-collapse:collapse; }}
        th, td {{ padding:10px; border-bottom:1px solid #e5e7eb; text-align:left; vertical-align:top; }}
        mark {{ background:#fde68a; padding:0 2px; }}
        .muted {{ color:#6b7280; }}
      </style>
    </head>
    <body>
      {nav}
      <div class='wrap'>
        {content}
      </div>
    </body>
    </html>
    """

    return HTMLResponse(html_doc)


@app.get("/favicon.ico")
async def favicon() -> Response:
    return Response(status_code=204)


@app.on_event("startup")
async def startup_event() -> None:
    init_db()


@app.get("/", response_class=HTMLResponse)
async def home() -> HTMLResponse:
    records = get_all_records()[:15]
    rows = ""

    for rec in records:
        analytics = rec.get("analytics_json") or {}
        timing = rec.get("timing_json") or {}

        rows += f"""
        <tr>
            <td><a href='/result/{rec['id']}'>{html.escape(rec['filename'])}</a></td>
            <td>{html.escape(rec.get('project_name') or '')}</td>
            <td>{rec.get('meeting_date') or ''}</td>
            <td>{round(rec.get('duration') or 0, 2)}</td>
            <td>{len(rec.get('tasks_json') or [])}</td>
            <td>{len(rec.get('qa_json') or [])}</td>
            <td>{analytics.get('avg_sentiment_score', 0)}</td>
            <td>{round(float((timing or {}).get('total', 0)), 2)}</td>
        </tr>
        """

    content = f"""
    <div class='section'>
      <h2>PM Insights</h2>
      <p class='muted'>Текущий режим NLP: <b>{html.escape(NLP_MODE)}</b></p>
      <form action='/upload/' method='post' enctype='multipart/form-data'>
        <div class='row'>
          <div style='flex:2'>
            <label>Аудиофайл</label>
            <input type='file' name='file' required>
          </div>
          <div style='flex:1'>
            <label>Дата встречи</label>
            <input type='date' name='meeting_date'>
          </div>
        </div>
        <div class='row' style='margin-top:10px'>
          <div style='flex:1'>
            <label>Проект</label>
            <input type='text' name='project_name' placeholder='Например, PM Insights'>
          </div>
          <div style='flex:1'>
            <label>Участники</label>
            <input type='text' name='participants' placeholder='Иван, Анна, Сергей'>
          </div>
        </div>
        <div style='margin-top:14px'>
          <button class='btn btn-blue' type='submit'>Загрузить и обработать</button>
        </div>
      </form>
    </div>
    <div class='section'>
      <h3>История обработанных встреч</h3>
      <div style='overflow-x:auto'>
        <table>
          <thead>
            <tr>
              <th>Файл</th>
              <th>Проект</th>
              <th>Дата</th>
              <th>Длительность</th>
              <th>Задачи</th>
              <th>Q/A</th>
              <th>Sentiment</th>
              <th>Общее время</th>
            </tr>
          </thead>
          <tbody>
            {rows or "<tr><td colspan='8' class='muted'>Пока нет данных</td></tr>"}
          </tbody>
        </table>
      </div>
    </div>
    """

    return page("PM Insights", content, "home")


@app.post("/upload/")
async def upload_file(
    file: UploadFile = File(...),
    meeting_date: Optional[str] = Form(None),
    project_name: str = Form(""),
    participants: str = Form(""),
):
    ext = Path(file.filename or "").suffix.lower()

    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Поддерживаются только: {', '.join(sorted(ALLOWED_EXTENSIONS))}")

    safe_name = Path(file.filename).name
    file_hash = hashlib.md5(f"{safe_name}_{time.time()}".encode()).hexdigest()[:12]

    original_path = UPLOAD_DIR / f"{file_hash}_{safe_name}"
    wav_path = CONVERTED_DIR / f"{file_hash}.wav"

    t0 = time.perf_counter()

    with original_path.open("wb") as f:
        shutil.copyfileobj(file.file, f)

    t1 = time.perf_counter()

    convert_to_wav(original_path, wav_path)

    t2 = time.perf_counter()

    transcription = transcribe_audio(wav_path)

    t3 = time.perf_counter()

    classified, sentiments, tasks, qa_pairs, analytics = process_segments(transcription["segments"])

    t4 = time.perf_counter()

    enriched_segments = []

    for seg, cls, sent in zip(transcription["segments"], classified, sentiments):
        enriched_segments.append({
            **seg,
            "predicted_label": cls.get("label", "other"),
            "prediction_confidence": cls.get("confidence", 0),
            "prediction_source": cls.get("source", "rules"),
            "prediction_debug": cls.get("debug", {}),
            "task_debug": cls.get("task_debug", {}),
            "sentiment_label": sent.get("label", "neutral"),
            "sentiment_score": sent.get("score", 0),
            "timecode": f"[{seg['start']:.2f} - {seg['end']:.2f} сек.]",
        })

    timing = {
        "upload": round(t1 - t0, 2),
        "convert": round(t2 - t1, 2),
        "transcribe": round(t3 - t2, 2),
        "nlp": round(t4 - t3, 2),
        "total": round(t4 - t0, 2),
    }

    payload = {
        "filename": safe_name,
        "project_name": project_name,
        "participants": participants,
        "meeting_date": meeting_date,
        "duration": transcription["duration"],
        "language": transcription["language"],
        "full_text": transcription["text"],
        "segments": enriched_segments,
        "tasks": tasks,
        "qa_pairs": qa_pairs,
        "analytics": analytics,
        "timing": timing,
    }

    stem = original_path.stem
    save_result_files(stem, payload)

    record_id = store_record(
        filename=safe_name,
        original_path=str(original_path),
        converted_path=str(wav_path),
        transcription={**transcription, "segments": enriched_segments},
        tasks=tasks,
        qa_pairs=qa_pairs,
        sentiments=sentiments,
        analytics=analytics,
        meeting_date=meeting_date,
        project_name=project_name,
        participants=participants,
        timing=timing,
    )

    current_record = get_record(record_id)
    previous_record = find_previous_record(record_id, project_name)

    if current_record:
        dynamic_analysis = build_dynamic_analysis(current_record, previous_record)
        analytics["dynamic_analysis"] = dynamic_analysis
        payload["analytics"] = analytics
        save_result_files(stem, payload)

        with get_conn() as conn, conn.cursor() as cur:
            cur.execute(
                """
                UPDATE transcriptions
                SET analytics_json = %s
                WHERE id = %s
                """,
                (Json(analytics), record_id),
            )
            conn.commit()

    return JSONResponse({
        "ok": True,
        "record_id": record_id,
        "redirect": f"/result/{record_id}",
    })


@app.get("/result/{record_id}", response_class=HTMLResponse)
async def result_page(record_id: int) -> HTMLResponse:
    rec = get_record(record_id)

    if not rec:
        raise HTTPException(404, "Запись не найдена")

    tasks = rec.get("tasks_json") or []
    qa_pairs = rec.get("qa_json") or []
    segments = rec.get("segments") or []
    analytics = rec.get("analytics_json") or {}
    timing = rec.get("timing_json") or {}

    dynamic = analytics.get("dynamic_analysis") or {}
    task_changes = dynamic.get("task_changes") or {}

    new_tasks = task_changes.get("new_tasks") or []
    repeated_tasks = task_changes.get("repeated_tasks") or []
    closed_tasks = task_changes.get("potentially_closed_tasks") or []
    aspect_changes = dynamic.get("aspect_changes") or []

    new_tasks_html = "".join(
        f"<div class='task-card'><b>{html.escape(t.get('text', ''))}</b>"
        f"<div class='sub'>Ответственный: {html.escape(t.get('responsible') or '—')} | "
        f"Срок: {html.escape(t.get('deadline') or '—')}</div></div>"
        for t in new_tasks
    )

    repeated_tasks_html = "".join(
        f"<div class='task-card'><b>Текущая:</b> {html.escape((item.get('current_task') or {}).get('text', ''))}<br>"
        f"<b>Похожая из прошлой встречи:</b> {html.escape((item.get('previous_task') or {}).get('text', ''))}"
        f"<div class='sub'>Сходство: {item.get('similarity', 0)}</div></div>"
        for item in repeated_tasks
    )

    closed_tasks_html = "".join(
        f"<div class='task-card'><b>{html.escape(t.get('text', ''))}</b>"
        f"<div class='sub'>Задача была в прошлой встрече, но не найдена в текущей.</div></div>"
        for t in closed_tasks
    )

    aspect_rows = "".join(
        f"<tr>"
        f"<td>{html.escape(a.get('aspect', ''))}</td>"
        f"<td>{a.get('previous_count', 0)}</td>"
        f"<td>{a.get('current_count', 0)}</td>"
        f"<td>{a.get('delta', 0)}</td>"
        f"</tr>"
        for a in aspect_changes
    )

    dynamic_html = f"""
    <div class='section'>
      <h3>Динамический анализ изменений</h3>
      <p>{html.escape(dynamic.get('summary', 'Нет данных для динамического анализа.'))}</p>

      <div class='cards'>
        <div class='card'><h4>Изменение задач</h4><div class='val'>{dynamic.get('tasks_delta', 0)}</div></div>
        <div class='card'><h4>Изменение Q/A</h4><div class='val'>{dynamic.get('qa_delta', 0)}</div></div>
        <div class='card'><h4>Изменение sentiment</h4><div class='val' style='font-size:22px'>{dynamic.get('sentiment_delta', 0)}</div></div>
        <div class='card'><h4>Изменение негатива</h4><div class='val' style='font-size:22px'>{dynamic.get('negative_ratio_delta', 0)}</div></div>
      </div>

      <h4>Новые задачи ({len(new_tasks)})</h4>
      {new_tasks_html or '<p class="muted">Новые задачи не найдены</p>'}

      <h4>Повторяющиеся задачи ({len(repeated_tasks)})</h4>
      {repeated_tasks_html or '<p class="muted">Повторяющиеся задачи не найдены</p>'}

      <h4>Потенциально закрытые задачи ({len(closed_tasks)})</h4>
      {closed_tasks_html or '<p class="muted">Потенциально закрытые задачи не найдены</p>'}

      <h4>Изменение проблемных аспектов</h4>
      <table>
        <thead>
          <tr>
            <th>Аспект</th>
            <th>Прошлая встреча</th>
            <th>Текущая встреча</th>
            <th>Изменение</th>
          </tr>
        </thead>
        <tbody>
          {aspect_rows or "<tr><td colspan='4' class='muted'>Нет данных</td></tr>"}
        </tbody>
      </table>
    </div>
    """

    tasks_html = "".join(
        f"<div class='task-card'><b>{i}. {html.escape(t.get('text', ''))}</b><div class='sub'>"
        f"Ответственный: <b>{html.escape(t.get('responsible') or '—')}</b> | "
        f"Срок: <b>{html.escape(t.get('deadline') or '—')}</b> | "
        f"Confidence: {t.get('confidence', 0)} | "
        f"Sentiment: {(t.get('sentiment') or {}).get('label', 'neutral')} | "
        f"{t.get('timecode', '')}</div>"
        f"<div class='sub'>Rule reasons: {html.escape(str((t.get('task_debug') or {}).get('reasons', [])))}</div></div>"
        for i, t in enumerate(tasks, 1)
    )

    qa_html = "".join(
        f"<div class='task-card'><b>В:</b> {html.escape(q.get('question', ''))} "
        f"<span class='muted'>{q.get('question_timecode', '')}</span><br>"
        f"<b>О:</b> {html.escape(q.get('answer', ''))} "
        f"<span class='muted'>{q.get('answer_timecode', '')}</span>"
        f"<div class='sub'>Confidence: {q.get('confidence', 0)}</div></div>"
        for q in qa_pairs
    )

    segments_html = ""
    chart_labels = []
    chart_scores = []

    for seg in segments:
        sent_label = seg.get("sentiment_label", "neutral")
        seg_class = "seg-neu"

        if sent_label == "positive":
            seg_class = "seg-pos"
        elif sent_label == "negative":
            seg_class = "seg-neg"

        chart_labels.append(seg.get("timecode", "")[:12])
        chart_scores.append(seg.get("sentiment_score", 0))

        segments_html += (
            f"<div class='seg {seg_class}'>"
            f"<span class='muted'>{seg.get('timecode', '')}</span> "
            f"{html.escape(seg.get('text', ''))}"
            f"<span style='float:right' class='muted'>{html.escape(sent_label)} ({seg.get('sentiment_score', 0)})</span>"
            f"<br><span class='muted'>pred: {html.escape(seg.get('predicted_label', 'other'))}, "
            f"conf={seg.get('prediction_confidence', 0)}, "
            f"src={html.escape(seg.get('prediction_source', 'rules'))}</span></div>"
        )

    content = f"""
    <div class='cards'>
      <div class='card'>
        <h4>Файл</h4>
        <div style='font-weight:700'>{html.escape(rec['filename'])}</div>
        <div class='sub'>Проект: {html.escape(rec.get('project_name') or '—')}</div>
      </div>
      <div class='card'>
        <h4>Дата</h4>
        <div class='val' style='font-size:22px'>{rec.get('meeting_date') or '—'}</div>
        <div class='sub'>Участники: {html.escape(rec.get('participants') or '—')}</div>
      </div>
      <div class='card'><h4>Длительность</h4><div class='val'>{rec.get('duration', 0)}<span style='font-size:14px'>с</span></div></div>
      <div class='card'><h4>Задач</h4><div class='val'>{len(tasks)}</div></div>
      <div class='card'><h4>Q/A пар</h4><div class='val'>{len(qa_pairs)}</div></div>
      <div class='card'>
        <h4>Sentiment</h4>
        <div class='val' style='font-size:20px'>{analytics.get('avg_sentiment_score', 0)}</div>
        <div class='sub'>Негатив: {analytics.get('negative_ratio', 0)}</div>
      </div>
    </div>

    <div class='timing'>
      <span>upload: {timing.get('upload', 0)}с</span>
      <span>convert: {timing.get('convert', 0)}с</span>
      <span>transcribe: {timing.get('transcribe', 0)}с</span>
      <span>nlp: {timing.get('nlp', 0)}с</span>
      <span>total: {timing.get('total', 0)}с</span>
    </div>

    <div class='row' style='margin:10px 0 18px'>
      <a class='btn btn-red' href='/export/pdf/{record_id}'>PDF</a>
      <a class='btn btn-green' href='/export/excel/{record_id}'>Excel</a>
      <a class='btn btn-blue' href='/results/{Path(rec['original_path']).stem}.json' target='_blank'>JSON</a>
      <a class='btn btn-gray' href='/'>← Назад</a>
    </div>

    {dynamic_html}

    <div class='section'><h3>Задачи ({len(tasks)})</h3>{tasks_html or '<p class="muted">Задачи не найдены</p>'}</div>
    <div class='section'><h3>Вопросы — ответы ({len(qa_pairs)})</h3>{qa_html or '<p class="muted">Q/A не найдены</p>'}</div>
    <div class='section'><h3>Транскрипция с тональностью</h3>{segments_html}</div>
    <div class='section'><h3>График тональности по сегментам</h3><canvas id='sentChart' height='180'></canvas></div>

    <script>
      new Chart(document.getElementById('sentChart'), {{
        type: 'bar',
        data: {{
          labels: {json.dumps(chart_labels)},
          datasets: [{{
            data: {json.dumps(chart_scores)},
            backgroundColor: '#3b82f6'
          }}]
        }},
        options: {{
          plugins: {{ legend: {{ display: false }} }},
          scales: {{ y: {{ min: -1, max: 1 }} }}
        }}
      }});
    </script>
    """

    return page(f"Результат — {rec['filename']}", content, "home")


@app.get("/analytics/", response_class=HTMLResponse)
async def analytics_page() -> HTMLResponse:
    records = list(reversed(get_all_records()))

    labels = [str(r.get("meeting_date") or str(r.get("created_at", ""))[:10]) for r in records]
    task_counts = [len(r.get("tasks_json") or []) for r in records]
    qa_counts = [len(r.get("qa_json") or []) for r in records]
    sentiments = [(r.get("analytics_json") or {}).get("avg_sentiment_score", 0) for r in records]
    negative_ratios = [(r.get("analytics_json") or {}).get("negative_ratio", 0) for r in records]

    total_tasks = sum(task_counts)
    total_qa = sum(qa_counts)
    avg_sent = round(sum(sentiments) / len(sentiments), 4) if sentiments else 0.0
    avg_negative = round(sum(negative_ratios) / len(negative_ratios), 4) if negative_ratios else 0.0

    rows = ""

    for r in reversed(records):
        analytics = r.get("analytics_json") or {}
        dynamic = analytics.get("dynamic_analysis") or {}

        rows += f"""
        <tr>
          <td><a href='/result/{r['id']}'>{html.escape(r['filename'])}</a></td>
          <td>{html.escape(r.get('project_name') or '')}</td>
          <td>{r.get('meeting_date') or ''}</td>
          <td>{len(r.get('tasks_json') or [])}</td>
          <td>{len(r.get('qa_json') or [])}</td>
          <td>{analytics.get('avg_sentiment_score', 0)}</td>
          <td>{analytics.get('negative_ratio', 0)}</td>
          <td>{dynamic.get('tasks_delta', 0)}</td>
          <td>{dynamic.get('sentiment_delta', 0)}</td>
        </tr>
        """

    content = f"""
    <div class='cards'>
      <div class='card'><h4>Встреч</h4><div class='val'>{len(records)}</div></div>
      <div class='card'><h4>Всего задач</h4><div class='val'>{total_tasks}</div></div>
      <div class='card'><h4>Всего Q/A</h4><div class='val'>{total_qa}</div></div>
      <div class='card'><h4>Ср. sentiment</h4><div class='val' style='font-size:22px'>{avg_sent}</div></div>
      <div class='card'><h4>Ср. негатив</h4><div class='val' style='font-size:22px'>{avg_negative}</div></div>
    </div>

    <div class='row'>
      <div class='section' style='flex:1'><h3>Динамика тональности</h3><canvas id='sentTrend'></canvas></div>
      <div class='section' style='flex:1'><h3>Задачи по встречам</h3><canvas id='taskTrend'></canvas></div>
    </div>

    <div class='row'>
      <div class='section' style='flex:1'><h3>Q/A по встречам</h3><canvas id='qaTrend'></canvas></div>
      <div class='section' style='flex:1'><h3>Доля негатива</h3><canvas id='negativeTrend'></canvas></div>
    </div>

    <div class='section'>
      <h3>Все встречи</h3>
      <div style='overflow-x:auto'>
        <table>
          <thead>
            <tr>
              <th>Файл</th>
              <th>Проект</th>
              <th>Дата</th>
              <th>Задачи</th>
              <th>Q/A</th>
              <th>Sentiment</th>
              <th>Негатив</th>
              <th>Δ задач</th>
              <th>Δ sentiment</th>
            </tr>
          </thead>
          <tbody>{rows or "<tr><td colspan='9' class='muted'>Нет данных</td></tr>"}</tbody>
        </table>
      </div>
    </div>

    <script>
      new Chart(document.getElementById('sentTrend'), {{
        type: 'line',
        data: {{
          labels: {json.dumps(labels)},
          datasets: [{{
            data: {json.dumps(sentiments)},
            label: 'Sentiment',
            borderColor: '#2563eb',
            fill: false
          }}]
        }},
        options: {{ scales: {{ y: {{ min: -1, max: 1 }} }} }}
      }});

      new Chart(document.getElementById('taskTrend'), {{
        type: 'bar',
        data: {{
          labels: {json.dumps(labels)},
          datasets: [{{
            data: {json.dumps(task_counts)},
            label: 'Tasks',
            backgroundColor: '#8b5cf6'
          }}]
        }},
        options: {{ plugins: {{ legend: {{ display: false }} }} }}
      }});

      new Chart(document.getElementById('qaTrend'), {{
        type: 'bar',
        data: {{
          labels: {json.dumps(labels)},
          datasets: [{{
            data: {json.dumps(qa_counts)},
            label: 'Q/A',
            backgroundColor: '#16a34a'
          }}]
        }},
        options: {{ plugins: {{ legend: {{ display: false }} }} }}
      }});

      new Chart(document.getElementById('negativeTrend'), {{
        type: 'line',
        data: {{
          labels: {json.dumps(labels)},
          datasets: [{{
            data: {json.dumps(negative_ratios)},
            label: 'Negative ratio',
            borderColor: '#dc2626',
            fill: false
          }}]
        }},
        options: {{ scales: {{ y: {{ min: 0, max: 1 }} }} }}
      }});
    </script>
    """

    return page("Аналитика", content, "analytics")


@app.get("/search/", response_class=HTMLResponse)
async def search_page(q: str = "") -> HTMLResponse:
    results = search_records(q) if q else []
    cards = ""

    for r in results:
        cards += f"""
        <div class='task-card'>
          <a href='/result/{r['id']}' style='font-weight:700'>{html.escape(r['filename'])}</a>
          <div class='sub'>{html.escape(r.get('project_name') or '')} | {r.get('meeting_date') or r.get('created_at') or ''}</div>
          <div style='margin-top:6px'>{r.get('snippet') or ''}</div>
        </div>
        """

    content = f"""
    <div class='section'>
      <h3>Поиск по встречам</h3>
      <form method='get' action='/search/'>
        <div class='row'>
          <div style='flex:1'>
            <input type='text' name='q' value='{html.escape(q)}' placeholder='Поиск по транскрипту, проекту, участникам'>
          </div>
          <div><button class='btn btn-blue' type='submit'>Найти</button></div>
        </div>
      </form>
      <div style='margin-top:12px'>
        {cards or ('<p class="muted">Введите запрос</p>' if not q else '<p class="muted">Ничего не найдено</p>')}
      </div>
    </div>
    """

    return page("Поиск", content, "search")


@app.get("/api/analytics", response_class=JSONResponse)
async def api_analytics() -> JSONResponse:
    records = get_all_records()

    payload = {
        "meetings_count": len(records),
        "total_tasks": sum(len(r.get("tasks_json") or []) for r in records),
        "total_qa_pairs": sum(len(r.get("qa_json") or []) for r in records),
        "avg_sentiment_overall": round(
            sum((r.get("analytics_json") or {}).get("avg_sentiment_score", 0) for r in records) / len(records),
            4,
        ) if records else 0,
        "avg_negative_ratio": round(
            sum((r.get("analytics_json") or {}).get("negative_ratio", 0) for r in records) / len(records),
            4,
        ) if records else 0,
    }

    return JSONResponse(payload)


@app.get("/export/excel/{record_id}")
async def excel_export(record_id: int):
    path = export_excel(record_id)

    return FileResponse(
        path,
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/export/pdf/{record_id}")
async def pdf_export(record_id: int):
    path = export_pdf(record_id)

    return FileResponse(
        path,
        filename=path.name,
        media_type="application/pdf",
    )


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "app:app",
        host="127.0.0.1",
        port=8000,
        reload=True,
    )